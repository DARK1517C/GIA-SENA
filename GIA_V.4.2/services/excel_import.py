from openpyxl import load_workbook
from .utils import normalize_header, clean_cell, build_alias_lookup

# --- Alias mappings (copiados del app.py original) ---
APPRENTICE_IMPORT_ALIASES = {
    "group_number": ["N° DE FICHA", "N DE FICHA", "NUMERO DE FICHA"],
    "lead_instructor": ["NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA", "NOMBRE DE INSTRUCTOR(A) LIDER DE LA FICHA", "INSTRUCTOR LIDER"],
    "followup_instructor": ["NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)", "INSTRUCTOR SEGUIMIENTO"],
    "followup_instructor_email": ["CORREO DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)", "CORREO INSTRUCTOR SEGUIMIENTO", "CORREO INSTRUCTOR ETAPA PRODUCTIVA"],
    "program_name": ["NOMBRE DEL PROGRAMA DE FORMACIÓN", "NOMBRE DEL PROGRAMA"],
    "document_type": ["TIPO DE DOCUMENTO (CC, TI, CE)", "TIPO DE DOCUMENTO"],
    "document_number": ["N° DE DOCUMENTO DEL APRENDIZ", "N DE DOCUMENTO DEL APRENDIZ", "NÚMERO DE DOCUMENTO"],
    "first_names": ["NOMBRES DEL APRENDIZ", "NOMBRES"],
    "last_names": ["APELLIDOS DEL APRENDIZ", "APELLIDOS"],
    "gender": ["GÉNERO (F/M)", "GENERO (F/M)", "GÉNERO"],
    "phone": ["TELÉFONO DEL APRENDIZ", "TELEFONO DEL APRENDIZ", "TELÉFONO"],
    "municipality_origin": ["MUNICIPIO DE ORIGEN"],
    "email": ["CORREO ELECTRÓNICO DEL APRENDIZ", "CORREO ELECTRONICO DEL APRENDIZ", "CORREO ELECTRÓNICO"],
    "ep_modality": ["MODALIDAD ETAPA PRODUCTIVA", "MODALIDAD EP"],
    "practice_start_date": ["FECHA INICIO DE PRÁCTICAS"],
    "practice_end_date": ["FECHA FINAL DE PRÁCTICAS"],
    "company_name": ["NOMBRE DE LA EMPRESA/ORG/INST", "NOMBRE EMPRESA"],
    "company_address": ["DIRECCIÓN DE LA EMPRESA", "DIRECCION DE LA EMPRESA", "DIRECCIÓN EMPRESA"],
    "company_municipality": ["MUNICIPIO", "MUNICIPIO EMPRESA"],
    "coformador_name": ["NOMBRE COFORMADOR"],
    "coformador_email": ["CORREO ELECTRÓNICO DEL COFORMADOR", "CORREO ELECTRONICO DEL COFORMADOR", "CORREO COFORMADOR"],
    "coformador_phone": ["TELÉFONO DEL COFORMADOR", "TELEFONO DEL COFORMADOR", "TELÉFONO COFORMADOR"],
    "individual_management": ["GESTIÓN INDIVIDUAL DEL APRENDIZ EN EP", "GESTION INDIVIDUAL DEL APRENDIZ EN EP", "GESTIÓN INDIVIDUAL"],
    "sofia_status": ["ESTADO DEL APRENDIZ EN SOFÍAPLUS", "ESTADO DEL APRENDIZ EN SOFIAPLUS", "ESTADO SOFIA PLUS"],
    "arl_responsible": ["RESPONSABLE DE AFILIACIÓN ARL", "RESPONSABLE DE AFILIACION ARL", "RESPONSABLE ARL"],
    "evaluation_date": [
        "FECHA EMISIÓN DE JUICIO EVALUATIVO EN SOFIA PLUS",
        "FECHA EMISION DE JUICIO EVALUATIVO EN SOFIA PLUS",
        "FECHA JUICIO EVALUATIVO",
    ],
    "english_results": ["JUICIOS DE INGLÉS APROBADOS SI/NO", "JUICIOS DE INGLES APROBADOS SI/NO", "JUICIOS DE INGLÉS"],
    "program_level": ["NIVEL DEL PROGRAMA", "NIVEL DE PROGRAMA"],
}

GROUP_IMPORT_ALIASES = {
    "group_number": ["N° DE FICHA", "N DE FICHA"],
    "lead_instructor": ["NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA", "NOMBRE DE INSTRUCTOR(A) LIDER DE LA FICHA"],
    "followup_instructor": ["NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)"],
    "program_name": ["NOMBRE DEL PROGRAMA DE FORMACIÓN"],
    "municipality": ["MUNICIPIO"],
    "program_level": ["NIVEL DE PROGRAMA"],
    "modality": ["MODALIDAD"],
    "sofia_group_status": ["ESTADO DE LA FICHA EN SOFÍAPLUS", "ESTADO DE LA FICHA EN SOFIAPLUS"],
    "group_start_date": ["FECHA INICIO DE LA FICHA EN SOFIAPLUS"],
    "training_end_date": ["FECHA FIN DE LA FORMACIÓN EN SOFIAPLUS", "FECHA FIN DE LA FORMACION EN SOFIAPLUS"],
    "ep_start_date": ["FECHA INICIO DE ETAPA PRODUCTIVA"],
    "group_validity": ["VIGENCIA DE LA FICHA"],
    "apprentices_training": ["APRENDICES EN FORMACIÓN", "APRENDICES EN FORMACION"],
    "apprentices_enabled": ["APRENDICES HABILITADOS PARA INICIAR ETAPA PRODUCTIVA", "APRENDICES HABILITADOS"],
    "apprentices_rap_pending": ["APRENDICES QUE DEBEN RAP", "APRENDICES DEBEN RAP"],
    "apprentices_practice": ["APRENDICES EN PRÁCTICA", "APRENDICES EN PRACTICA"],
    "apprentices_without_alternative": ["APRENDICES SIN ALTERNATIVA DE PRÁCTIVA", "APRENDICES SIN ALTERNATIVA DE PRACTIVA"],
    "apprentices_certified": ["APRENDICES CERTIFICADOS"],
    "learning_contract": ["CONTRATO DE APRENDIZAJE", "CONTRATO APRENDIZAJE"],
    "internship": ["PASANTIA", "PASANTÍA"],
    "productive_project": ["PROYECTO PRODUCTIVO"],
    "employment_link": ["VINCULACION LABORAL", "VINCULACIÓN LABORAL"],
}

APPRENTICE_MODEL_FIELDS = [
    "group_number", "document_type", "document_number", "first_names", "last_names",
    "gender", "phone", "email", "municipality_origin", "program_name", "group_validity",
    "lead_instructor", "followup_instructor", "followup_instructor_email", "ep_modality", "sofia_status",
    "practice_start_date", "practice_end_date", "company_name", "company_municipality",
    "company_address", "coformador_name", "coformador_email", "coformador_phone",
    "arl_responsible", "individual_management", "followup_moments", "evaluation_date",
    "english_results", "program_level"
]

GROUP_MODEL_FIELDS = [
    "group_number", "program_name", "lead_instructor", "followup_instructor",
    "municipality", "program_level", "modality", "sofia_group_status", "group_validity",
    "group_start_date", "training_end_date", "ep_start_date", "apprentices_statistics",
    "apprentices_training", "apprentices_enabled", "apprentices_rap_pending",
    "apprentices_practice", "apprentices_without_alternative", "apprentices_certified",
    "productive_modalities", "learning_contract", "internship", "productive_project",
    "employment_link"
]

APPRENTICE_ALIAS_LOOKUP = build_alias_lookup(APPRENTICE_IMPORT_ALIASES)
GROUP_ALIAS_LOOKUP = build_alias_lookup(GROUP_IMPORT_ALIASES)


def find_sheet_by_headers(workbook, required_headers, alias_lookup, min_matches=4):
    best_match = None
    best_score = 0
    required = {normalize_header(item) for item in required_headers}
    for sheet in workbook.worksheets:
        for row_index in range(1, min(sheet.max_row, 8) + 1):
            values = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
            normalized = [normalize_header(value) for value in values if normalize_header(value)]
            matches = sum(1 for item in normalized if item in required or item in alias_lookup)
            if matches > best_score:
                best_score = matches
                best_match = (sheet, row_index)
    if best_match and best_score >= min_matches:
        return best_match
    return None, None


def extract_sheet_rows(sheet, header_row):
    headers = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
    data = []
    for row_index in range(header_row + 1, sheet.max_row + 1):
        row_values = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "", "\xa0") for value in row_values):
            continue
        data.append(row_values)
    return headers, data


def parse_apprentice_sheet(sheet, header_row):
    headers, data_rows = extract_sheet_rows(sheet, header_row)
    header_map = {index + 1: normalize_header(value) for index, value in enumerate(headers) if normalize_header(value)}
    records = {}
    for row in data_rows:
        record = {key: "" for key in APPRENTICE_MODEL_FIELDS}
        for col_index, normalized in header_map.items():
            field = APPRENTICE_ALIAS_LOOKUP.get(normalized)
            if not field:
                continue
            value = clean_cell(row[col_index - 1] if col_index - 1 < len(row) else "")
            if field in record:
                record[field] = value
        moments = []
        for col_index in range(17, 21):
            if col_index - 1 < len(row):
                value = clean_cell(row[col_index - 1])
                if value:
                    moments.append(value)
        if moments:
            record["followup_moments"] = " | ".join(moments)
        if record.get("document_number"):
            records[record["document_number"]] = record
    return list(records.values())


def parse_group_sheet(sheet, header_row):
    headers = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
    subheaders = [sheet.cell(header_row + 1, col).value for col in range(1, sheet.max_column + 1)] if sheet.max_row > header_row else []
    header_map = {}
    for index, value in enumerate(headers, start=1):
        normalized = normalize_header(value)
        if normalized:
            header_map[index] = normalized
    for index, value in enumerate(subheaders, start=1):
        normalized = normalize_header(value)
        if normalized:
            header_map[index] = normalized
    start_data_row = header_row + 2 if any(normalize_header(item) for item in subheaders) else header_row + 1
    records = {}
    for row_index in range(start_data_row, sheet.max_row + 1):
        row = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "", "\xa0") for value in row):
            continue
        record = {key: "" for key in GROUP_MODEL_FIELDS}
        for col_index, normalized in header_map.items():
            field = GROUP_ALIAS_LOOKUP.get(normalized)
            if not field:
                continue
            record[field] = clean_cell(row[col_index - 1] if col_index - 1 < len(row) else "")
        modality_parts = []
        for key, label in [
            ("learning_contract", "Contrato aprendizaje"),
            ("internship", "Pasantía"),
            ("productive_project", "Proyecto productivo"),
            ("employment_link", "Vinculación laboral"),
        ]:
            if record.get(key):
                modality_parts.append(f"{label}: {record[key]}")
        record["productive_modalities"] = " | ".join(modality_parts)
        total_stats = [
            record.get("apprentices_training"),
            record.get("apprentices_enabled"),
            record.get("apprentices_rap_pending"),
            record.get("apprentices_practice"),
            record.get("apprentices_without_alternative"),
            record.get("apprentices_certified"),
        ]
        record["apprentices_statistics"] = " / ".join(value for value in total_stats if value)
        if record.get("group_number"):
            records[record["group_number"]] = record
    return list(records.values())


def upsert_student_user(apprentice, known_users=None):
    """
    Crea o actualiza usuario tipo 'aprendiz' asociado al aprendiz si aplica.
    known_users: dict username->User para evitar consultas repetidas.
    """
    from models import User
    from extensions import db

    if known_users is None:
        known_users = {u.username: u for u in User.query.filter_by(role="aprendiz").all()}

    # apprentice puede ser dict (antes de persistir) o instancia ORM
    if isinstance(apprentice, dict):
        username = apprentice.get("document_number")
        first_names = apprentice.get("first_names", "")
        last_names = apprentice.get("last_names", "")
        email = apprentice.get("email")
    else:
        username = getattr(apprentice, "document_number", None)
        first_names = getattr(apprentice, "first_names", "")
        last_names = getattr(apprentice, "last_names", "")
        email = getattr(apprentice, "email", None)

    if not username:
        return None

    user = known_users.get(username) or User.query.filter_by(username=username).first()
    if user is None:
        user = User(username=username, full_name=f"{first_names} {last_names}".strip(), role="aprendiz", email=email)
        user.set_password(username)
        db.session.add(user)
        db.session.flush()
        known_users[username] = user
    else:
        user.full_name = f"{first_names} {last_names}".strip() or user.full_name
        if email:
            user.email = email

    if isinstance(apprentice, dict):
        return user.id
    else:
        apprentice.student_user_id = user.id
        return user.id


def import_reference_workbook(file_storage, owner_id, mode="both"):
    """
    Importa hojas de referencia desde un archivo .xlsx.
    Retorna: (apprentice_count, group_count, has_apprentice_sheet, has_group_sheet)
    """
    from models import Apprentice, TrainingGroup, User
    from extensions import db
    from services.evidence_service import ensure_submissions_for_apprentice, seed_default_evidences_for_group

    workbook = load_workbook(file_storage, data_only=False)
    apprentice_sheet = apprentice_header = None
    group_sheet = group_header = None

    if mode in {"both", "apprentices"}:
        apprentice_sheet, apprentice_header = find_sheet_by_headers(
            workbook,
            [
                "N° DE DOCUMENTO DEL APRENDIZ",
                "NOMBRES DEL APRENDIZ",
                "MODALIDAD ETAPA PRODUCTIVA",
                "GESTIÓN INDIVIDUAL DEL APRENDIZ EN EP",
            ],
            APPRENTICE_ALIAS_LOOKUP,
        )

    if mode in {"both", "groups"}:
        group_sheet, group_header = find_sheet_by_headers(
            workbook,
            [
                "N° DE FICHA",
                "APRENDICES EN FORMACIÓN",
                "APRENDICES EN PRÁCTICA",
                "APRENDICES CERTIFICADOS",
            ],
            GROUP_ALIAS_LOOKUP,
        )

    apprentice_count = 0
    group_count = 0
    existing_apprentices = {item.document_number: item for item in Apprentice.query.all()}
    existing_groups = {item.group_number: item for item in TrainingGroup.query.all()}
    known_users = {item.username: item for item in User.query.filter_by(role="aprendiz").all()}

    if apprentice_sheet is not None:
        for index, data in enumerate(parse_apprentice_sheet(apprentice_sheet, apprentice_header), start=1):
            clean_data = {key: value for key, value in data.items() if key in APPRENTICE_MODEL_FIELDS}
            apprentice_obj = existing_apprentices.get(data["document_number"])
            if apprentice_obj is None:
                apprentice_obj = Apprentice(created_by=owner_id, **clean_data)
                db.session.add(apprentice_obj)
                db.session.flush()
                existing_apprentices[data["document_number"]] = apprentice_obj
            else:
                for key, value in clean_data.items():
                    if hasattr(apprentice_obj, key):
                        setattr(apprentice_obj, key, value)
                apprentice_obj.created_by = owner_id
            upsert_student_user(apprentice_obj, known_users=known_users)
            group_number = clean_data.get("group_number")
            if group_number and group_number not in existing_groups:
                group_obj = TrainingGroup(
                    created_by=owner_id,
                    group_number=group_number,
                    program_name=clean_data.get("program_name") or "Sin programa",
                    lead_instructor=clean_data.get("lead_instructor"),
                    followup_instructor=clean_data.get("followup_instructor"),
                    municipality=clean_data.get("company_municipality") or clean_data.get("municipality_origin"),
                    program_level=clean_data.get("program_level"),
                    group_validity=clean_data.get("group_validity"),
                )
                db.session.add(group_obj)
                db.session.flush()
                existing_groups[group_number] = group_obj
                group_count += 1
                seed_default_evidences_for_group(group_obj)
            elif group_number:
                seed_default_evidences_for_group(existing_groups[group_number])
            ensure_submissions_for_apprentice(apprentice_obj)
            apprentice_count += 1
            if index % 100 == 0:
                db.session.commit()

    if group_sheet is not None:
        for index, data in enumerate(parse_group_sheet(group_sheet, group_header), start=1):
            clean_data = {key: value for key, value in data.items() if key in GROUP_MODEL_FIELDS}
            group_obj = existing_groups.get(data["group_number"])
            if group_obj is None:
                group_obj = TrainingGroup(created_by=owner_id, **clean_data)
                db.session.add(group_obj)
                db.session.flush()
                existing_groups[data["group_number"]] = group_obj
            else:
                for key, value in clean_data.items():
                    if hasattr(group_obj, key):
                        setattr(group_obj, key, value)
                group_obj.created_by = owner_id
            seed_default_evidences_for_group(group_obj)
            group_count += 1
            if index % 100 == 0:
                db.session.commit()

    db.session.commit()
    return apprentice_count, group_count, apprentice_sheet is not None, group_sheet is not None
