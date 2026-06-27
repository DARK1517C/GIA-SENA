# services/excel_export.py
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.group import TrainingGroup
from services.utils import followup_range_label


def format_value(value):
    try:
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        return str(value)
    except Exception:
        return str(value)


def export_workbook(title, fields, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    headers = [label for _key, label in fields]
    sheet.append(headers)
    for column_index, _header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B8F47")
    for row in rows:
        sheet.append([
            format_value(row.get(key, "") if isinstance(row, dict) else getattr(row, key, ""))
            for key, _label in fields
        ])
    _autowidth(workbook)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def write_template_headers(sheet, top_headers, sub_headers=None):
    green_fill = PatternFill("solid", fgColor="0B8F47")
    white_font = Font(bold=True, color="FFFFFF")
    for index, value in enumerate(top_headers, start=1):
        cell = sheet.cell(1, index)
        cell.value = value
        cell.fill = green_fill
        cell.font = white_font
    if sub_headers:
        for index, value in enumerate(sub_headers, start=1):
            cell = sheet.cell(2, index)
            cell.value = value
            cell.fill = green_fill
            cell.font = white_font


def _get(row, key):
    return row.get(key, "") if isinstance(row, dict) else getattr(row, key, "")


def _autowidth(workbook):
    for sheet in workbook.worksheets:
        for col_idx, column in enumerate(sheet.columns, start=1):
            max_length = 0
            for cell in column:
                text = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(text))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 38)


def export_reference_workbook(apprentice_rows, group_rows):
    workbook = Workbook()
    apprentice_sheet = workbook.active
    apprentice_sheet.title = "Aprendices"

    apprentice_headers = [
        "CONSECUTIVO",
        "N° DE FICHA",
        "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
        "NOMBRE DE INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "NOMBRE DEL PROGRAMA DE FORMACIÓN",
        "NIVEL DEL PROGRAMA",
        "TIPO DE DOCUMENTO (CC, TI, CE)",
        "N° DOCUMENTO DEL APRENDIZ",
        "NOMBRES DEL APRENDIZ",
        "APELLIDOS DEL APRENDIZ",
        "GÉNERO (F/M)",
        "N° DE CONTACTO DEL APRENDIZ",
        "MUNICIPIO DE ORIGEN",
        "CORREO ELECTRÓNICO DEL APRENDIZ",
        "ALTERNATIVA ETAPA PRODUCTIVA",
        "FECHA INICIO DE PRÁCTICAS",
        "FECHA FINAL DE PRÁCTICAS",
        "MOMENTO 1 - RANGO",
        "MOMENTO 2 - RANGO",
        "MOMENTO 3 - RANGO",
        "MOMENTO 4 - RANGO",
        "NOMBRE DE LA EMPRESA/ORG/INST",
        "DIRECCIÓN DE LA EMPRESA",
        "MUNICIPIO",
        "NOMBRE COFORMADOR",
        "CORREO ELECTRÓNICO DEL COFORMADOR",
        "TELÉFONO DEL COFORMADOR",
        "GESTIÓN INDIVIDUAL DEL APRENDIZ",
        "ESTADO DEL APRENDIZ EN SOFÍAPLUS",
        "RESPONSABLE DE AFILIACIÓN ARL",
        "CONTINÚA EN LA EMPRESA (Si/No)",
    ]
    write_template_headers(apprentice_sheet, apprentice_headers)

    for index, apprentice in enumerate(apprentice_rows, start=1):
        group = _get(apprentice, "group")
        group_number = _get(group, "group_number") if group else _get(apprentice, "group_number")
        row = [
            index,
            group_number,
            _get(apprentice, "lead_instructor"),
            _get(apprentice, "followup_instructor"),
            _get(apprentice, "program_name"),
            _get(apprentice, "program_level"),
            _get(apprentice, "document_type"),
            _get(apprentice, "document_number"),
            _get(apprentice, "first_names"),
            _get(apprentice, "last_names"),
            _get(apprentice, "gender"),
            _get(apprentice, "phone"),
            _get(apprentice, "municipality_origin"),
            _get(apprentice, "email"),
            _get(apprentice, "ep_modality"),
            _get(apprentice, "practice_start_date"),
            _get(apprentice, "practice_end_date"),
            followup_range_label(_get(apprentice, "followup_moment1_start"), _get(apprentice, "followup_moment1_end")),
            followup_range_label(_get(apprentice, "followup_moment2_start"), _get(apprentice, "followup_moment2_end")),
            followup_range_label(_get(apprentice, "followup_moment3_start"), _get(apprentice, "followup_moment3_end")),
            followup_range_label(_get(apprentice, "followup_moment4_start"), _get(apprentice, "followup_moment4_end")),
            _get(apprentice, "company_name"),
            _get(apprentice, "company_address"),
            _get(apprentice, "company_municipality"),
            _get(apprentice, "coformador_name"),
            _get(apprentice, "coformador_email"),
            _get(apprentice, "coformador_phone"),
            _get(apprentice, "individual_management"),
            _get(apprentice, "sofia_status"),
            _get(apprentice, "arl_responsible"),
            _get(apprentice, "continues_company"),
        ]
        apprentice_sheet.append([format_value(value) for value in row])

    group_sheet = workbook.create_sheet("Record Fichas")
    group_headers = [
        "OBSERVACIÓN",
        "N° DE FICHA",
        "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
        "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)",
        "NOMBRE DEL PROGRAMA DE FORMACIÓN",
        "MUNICIPIO",
        "NIVEL DE PROGRAMA",
        "MODALIDAD",
        "ESTADO DE LA FICHA EN SOFÍAPLUS",
        "FECHA INICIO DE LA FICHA EN SOFIAPLUS",
        "FECHA FIN DE LA FORMACIÓN EN SOFIAPLUS",
        "FECHA INICIO DE ETAPA PRODUCTIVA",
        "VIGENCIA DE LA FICHA",
        "APRENDICES EN FORMACIÓN",
        "APRENDICES HABILITADOS PARA INICIARETAPA PRODUCTIVA",
        "APRENDICES QUE DEBEN RAP",
        "APRENDICES EN PRÁCTICA",
        None,
        None,
        None,
        "APRENDICES CERTIFICADOS",
        "TOTAL APRENDICES RELACIONADOS",
    ]
    group_subheaders = [None] * 16 + ["Ccontrato de aprendizaje", "Vínculo Formativo", "Vinculación Laboral", "Proyecto Productivo", None, None]
    write_template_headers(group_sheet, group_headers, group_subheaders)
    for group in group_rows:
        group_sheet.append([
            "",
            _get(group, "group_number"),
            _get(group, "lead_instructor"),
            _get(group, "followup_instructor"),
            _get(group, "program_name"),
            _get(group, "municipality"),
            _get(group, "program_level"),
            _get(group, "modality"),
            _get(group, "sofia_group_status"),
            _get(group, "group_start_date"),
            _get(group, "training_end_date"),
            _get(group, "ep_start_date"),
            _get(group, "group_validity"),
            _get(group, "apprentices_training"),
            _get(group, "apprentices_enabled"),
            _get(group, "apprentices_rap_pending"),
            _get(group, "apprentices_practice"),
            _get(group, "learning_contract"),
            _get(group, "internship"),
            _get(group, "employment_link"),
            _get(group, "productive_project"),
            len(getattr(group, "apprentices", []) or []),
        ])

    _autowidth(workbook)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
