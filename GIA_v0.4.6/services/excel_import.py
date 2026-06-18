from dataclasses import dataclass, field
from openpyxl import load_workbook
from .utils import (
    normalize_header,
    clean_cell,
    build_alias_lookup,
    calculate_followup_ranges,
    calculate_group_validity,
    validate_group_validity,
)

OFFICIAL_GROUP_SHEET = "Record Fichas"
OFFICIAL_APPRENTICE_SHEET = "Aprendices"


@dataclass
class ImportResult:
    apprentice_count: int = 0
    group_count: int = 0
    has_apprentice_sheet: bool = False
    has_group_sheet: bool = False
    created_apprentices: int = 0
    updated_apprentices: int = 0
    created_groups: int = 0
    updated_groups: int = 0
    skipped_apprentices: int = 0
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def __iter__(self):
        yield self.apprentice_count
        yield self.group_count
        yield self.has_apprentice_sheet
        yield self.has_group_sheet


APPRENTICE_IMPORT_ALIASES = {
    "group_number": ["N° DE FICHA", "Nº DE FICHA", "N DE FICHA", "FICHA"],
    "lead_instructor": ["NOMBRE DE INSTRUCTOR(A) LIDER DE LA FICHA", "NOMBRE DEL INSTRUCTOR LIDER DE LA FICHA"],
    "followup_instructor": ["NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)", "NOMBRE DE INSTRUCTOR(A) ETAPA PRODUCTIVA", "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO EP"],
    "followup_instructor_email": ["CORREO DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)", "CORREO INSTRUCTOR SEGUIMIENTO"],
    "program_name": ["NOMBRE DEL PROGRAMA DE FORMACION", "NOMBRE DEL PROGRAMA DE FORMACON"],
    "program_level": ["NIVEL DEL PROGRAMA", "NIVEL DE PROGRAMA"],
    "document_type": ["TIPO DE DOCUMENTO (CC, TI, CE)", "TIPO DE DOCUMENTO"],
    "document_number": ["N° DOCUMENTO DEL APRENDIZ", "N° DE DOCUMENTO DEL APRENDIZ", "N° DOCUMENTO DE IDENTIFICACION", "Nº DOCUMENTO DE IDENTIFICACION"],
    "first_names": ["NOMBRES DEL APRENDIZ", "NOMBRES"],
    "last_names": ["APELLIDOS DEL APRENDIZ", "APELLIDOS"],
    "gender": ["GENERO (F/M)", "GENERO"],
    "phone": ["N° DE CONTACTO DEL APRENDIZ", "TELEFONO DEL APRENDIZ", "TELEFONO"],
    "municipality_origin": ["MUNICIPIO DE ORIGEN", "CIUDAD DE ORIGEN DEL APRENDIZ"],
    "email": ["CORREO ELECTRONICO DEL APRENDIZ", "CORREO ELECTRONICO"],
    "ep_modality": ["ALTERNATIVA ETAPA PRODUCTIVA", "MODALIDAD ETAPA PRODUCTIVA", "ALTERNATIVA EP"],
    "practice_start_date": ["FECHA INICIO DE PRACTICAS", "FECHA DE INICIO DE EP"],
    "practice_end_date": ["FECHA FINAL DE PRACTICAS", "FECHA FINAL DE EP"],
    "company_name": ["NOMBRE DE LA EMPRESA/ORG/INST", "NOMBRE EMPRESA"],
    "company_address": ["DIRECCION DE LA EMPRESA", "DIRECCION DE EMPRESA"],
    "company_municipality": ["MUNICIPIO", "MUNICIPIO EMPRESA"],
    "coformador_name": ["NOMBRE COFORMADOR", "NOMBRE DEL COFORMADOR"],
    "coformador_email": ["CORREO ELECTRONICO DEL COFORMADOR", "CORREO DEL COFORMADOR"],
    "coformador_phone": ["TELEFONO DEL COFORMADOR"],
    "individual_management": ["GESTION INDIVIDUAL DEL APRENDIZ", "GESTION INDIVIDUAL DEL APRENDIZ EN EP"],
    "sofia_status": ["ESTADO DEL APRENDIZ EN SOFIAPLUS", "ESTADO SOFIA PLUS"],
    "arl_responsible": ["RESPONSABLE DE AFILIACION ARL", "RESPONSABLE ARL"],
    "continues_company": ["CONTINUA EN LA EMPRESA (SI/NO)", "SE QUEDA LABORANDO EN LA EMPRESA"],
    "evaluation_date": ["FECHA EMISION DE JUICIO EVALUATIVO EN SOFIA PLUS", "FECHA DE EVALUACION DE ETAPA PRODUCTIVA EN SOFIA PLUS"],
    "english_results": ["JUICIOS DE INGLES APROBADOS SI/NO", "JUICIOS DE INGLES"],
}

GROUP_IMPORT_ALIASES = {
    "group_number": ["N° DE FICHA", "Nº DE FICHA", "N DE FICHA", "FICHA"],
    "lead_instructor": ["NOMBRE DE INSTRUCTOR(A) LIDER DE LA FICHA", "NOMBRE DEL INSTRUCTOR LIDER DE LA FICHA"],
    "followup_instructor": ["NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)", "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO EP"],
    "program_name": ["NOMBRE DEL PROGRAMA DE FORMACION"],
    "municipality": ["MUNICIPIO"],
    "program_level": ["NIVEL DE PROGRAMA", "NIVEL DEL PROGRAMA"],
    "modality": ["MODALIDAD"],
    "sofia_group_status": ["ESTADO DE LA FICHA EN SOFIAPLUS"],
    "group_start_date": ["FECHA INICIO DE LA FICHA EN SOFIAPLUS"],
    "training_end_date": ["FECHA FIN DE LA FORMACION EN SOFIAPLUS"],
    "ep_start_date": ["FECHA INICIO DE ETAPA PRODUCTIVA"],
    "group_validity": ["VIGENCIA DE LA FICHA"],
    "apprentices_training": ["APRENDICES EN FORMACION"],
    "apprentices_enabled": ["APRENDICES HABILITADOS PARA INICIARETAPA PRODUCTIVA", "APRENDICES HABILITADOS PARA INICIAR ETAPA PRODUCTIVA"],
    "apprentices_rap_pending": ["APRENDICES QUE DEBEN RAP"],
    "apprentices_practice": ["APRENDICES EN PRACTICA"],
    "apprentices_certified": ["APRENDICES CERTIFICADOS"],
    "learning_contract": ["CCONTRATO DE APRENDIZAJE", "CONTRATO DE APRENDIZAJE"],
    "internship": ["VINCULO FORMATIVO", "PASANTIA", "PASANTIA"],
    "employment_link": ["VINCULACION LABORAL"],
    "productive_project": ["PROYECTO PRODUCTIVO"],
}

APPRENTICE_MODEL_FIELDS = [
    "group_id", "group_number", "document_type", "document_number", "first_names", "last_names",
    "gender", "phone", "email", "municipality_origin", "program_name", "group_validity",
    "lead_instructor", "followup_instructor", "followup_instructor_email", "ep_modality", "sofia_status",
    "practice_start_date", "practice_end_date", "followup_moment1_start", "followup_moment1_end",
    "followup_moment2_start", "followup_moment2_end", "followup_moment3_start", "followup_moment3_end",
    "followup_moment4_start", "followup_moment4_end", "company_name", "company_municipality",
    "company_address", "coformador_name", "coformador_email", "coformador_phone",
    "arl_responsible", "continues_company", "individual_management", "followup_moments", "evaluation_date",
    "english_results", "program_level"
]

GROUP_MODEL_FIELDS = [
    "group_number", "program_name", "lead_instructor", "followup_instructor",
    "municipality", "program_level", "modality", "sofia_group_status", "group_validity",
    "group_start_date", "training_end_date", "ep_start_date", "apprentices_statistics",
    "apprentices_training", "apprentices_enabled", "apprentices_rap_pending",
    "apprentices_practice", "apprentices_without_alternative", "apprentices_certified",
    "productive_modalities", "learning_contract", "internship", "productive_project",
    "employment_link"
]

APPRENTICE_ALIAS_LOOKUP = build_alias_lookup(APPRENTICE_IMPORT_ALIASES)
GROUP_ALIAS_LOOKUP = build_alias_lookup(GROUP_IMPORT_ALIASES)


def _sheet_by_official_name(workbook, official_name):
    wanted = normalize_header(official_name)
    for sheet in workbook.worksheets:
        if normalize_header(sheet.title) == wanted:
            return sheet
    return None


def _header_map(sheet, header_row, subheader_row=None, alias_lookup=None):
    mapping = {}
    alias_lookup = alias_lookup or {}
    for col in range(1, sheet.max_column + 1):
        candidates = [sheet.cell(header_row, col).value]
        if subheader_row:
            candidates.insert(0, sheet.cell(subheader_row, col).value)
        for value in candidates:
            normalized = normalize_header(value)
            if normalized and normalized in alias_lookup:
                mapping[col] = alias_lookup[normalized]
                break
    return mapping


def parse_apprentice_sheet(sheet):
    header_map = _header_map(sheet, 1, alias_lookup=APPRENTICE_ALIAS_LOOKUP)
    records = {}
    for row_index in range(2, sheet.max_row + 1):
        record = {key: "" for key in APPRENTICE_MODEL_FIELDS}
        has_data = False
        for col_index, field in header_map.items():
            value = clean_cell(sheet.cell(row_index, col_index).value)
            if value:
                has_data = True
            if field in record:
                record[field] = value
        moments = []
        for col_index in range(18, 22):
            value = clean_cell(sheet.cell(row_index, col_index).value)
            if value:
                moments.append(value)
        while len(moments) < 4:
            moments.append("")
        record["followup_moments"] = " | ".join(moments).strip(" |")
        for key, value in calculate_followup_ranges(record.get("practice_start_date"), record.get("practice_end_date")).items():
            if not record.get(key):
                record[key] = value
        if has_data and record.get("document_number"):
            records[record["document_number"]] = record
    return list(records.values())


def parse_group_sheet(sheet, result=None):
    header_map = _header_map(sheet, 2, subheader_row=3, alias_lookup=GROUP_ALIAS_LOOKUP)
    records = {}
    for row_index in range(4, sheet.max_row + 1):
        record = {key: "" for key in GROUP_MODEL_FIELDS}
        has_data = False
        for col_index, field in header_map.items():
            value = clean_cell(sheet.cell(row_index, col_index).value)
            if value:
                has_data = True
            if field in record:
                record[field] = value
        if not record.get("group_validity"):
            record["group_validity"] = calculate_group_validity(record.get("training_end_date"))
        notes = validate_group_validity(record.get("ep_start_date"), record.get("training_end_date"), record.get("group_validity"))
        if notes and result:
            result.warnings.append(f"Ficha {record.get('group_number')}: " + " ".join(notes))
        modality_parts = []
        for key, label in [
            ("learning_contract", "Contrato de aprendizaje"),
            ("internship", "Vinculo formativo"),
            ("employment_link", "Vinculacion laboral"),
            ("productive_project", "Proyecto productivo"),
        ]:
            if record.get(key):
                modality_parts.append(f"{label}: {record[key]}")
        record["productive_modalities"] = " | ".join(modality_parts)
        total_stats = [
            record.get("apprentices_training"),
            record.get("apprentices_enabled"),
            record.get("apprentices_rap_pending"),
            record.get("apprentices_practice"),
            record.get("apprentices_certified"),
        ]
        record["apprentices_statistics"] = " / ".join(value for value in total_stats if value)
        if has_data and record.get("group_number"):
            records[record["group_number"]] = record
    return list(records.values())


def upsert_student_user(apprentice, known_users=None):
    from models import User
    from extensions import db

    if known_users is None:
        known_users = {u.username: u for u in User.query.filter_by(role="aprendiz").all()}

    username = apprentice.get("document_number") if isinstance(apprentice, dict) else getattr(apprentice, "document_number", None)
    first_names = apprentice.get("first_names", "") if isinstance(apprentice, dict) else getattr(apprentice, "first_names", "")
    last_names = apprentice.get("last_names", "") if isinstance(apprentice, dict) else getattr(apprentice, "last_names", "")
    email = apprentice.get("email") if isinstance(apprentice, dict) else getattr(apprentice, "email", None)
    if not username:
        return None
    user = known_users.get(username) or User.query.filter_by(username=username).first()
    if user is None:
        user = User(username=username, full_name=f"{first_names} {last_names}".strip(), role="aprendiz", email=email)
        user.set_password(username)
        db.session.add(user)
        db.session.flush()
        known_users[username] = user
    else:
        user.full_name = f"{first_names} {last_names}".strip() or user.full_name
        if email:
            user.email = email
    if isinstance(apprentice, dict):
        return user.id
    apprentice.student_user_id = user.id
    return user.id


def import_reference_workbook(file_storage, owner_id, mode="both"):
    from models import Apprentice, TrainingGroup, User
    from extensions import db
    from services.evidence_service import ensure_submissions_for_apprentice, seed_default_evidences_for_group, sync_group_followup_dates

    result = ImportResult()
    workbook = load_workbook(file_storage, data_only=False)
    apprentice_sheet = _sheet_by_official_name(workbook, OFFICIAL_APPRENTICE_SHEET) if mode in {"both", "apprentices"} else None
    group_sheet = _sheet_by_official_name(workbook, OFFICIAL_GROUP_SHEET) if mode in {"both", "groups"} else None
    result.has_apprentice_sheet = apprentice_sheet is not None
    result.has_group_sheet = group_sheet is not None

    existing_apprentices = {item.document_number: item for item in Apprentice.query.all()}
    existing_groups = {item.group_number: item for item in TrainingGroup.query.all()}
    known_users = {item.username: item for item in User.query.filter_by(role="aprendiz").all()}

    if group_sheet is not None:
        for index, data in enumerate(parse_group_sheet(group_sheet, result), start=1):
            clean_data = {key: value for key, value in data.items() if key in GROUP_MODEL_FIELDS}
            group_obj = existing_groups.get(data["group_number"])
            if group_obj is None:
                group_obj = TrainingGroup(created_by=owner_id, **clean_data)
                db.session.add(group_obj)
                db.session.flush()
                existing_groups[data["group_number"]] = group_obj
                result.created_groups += 1
            else:
                for key, value in clean_data.items():
                    if hasattr(group_obj, key):
                        setattr(group_obj, key, value)
                group_obj.created_by = owner_id
                result.updated_groups += 1
            seed_default_evidences_for_group(group_obj)
            sync_group_followup_dates(group_obj)
            result.group_count += 1
            if index % 100 == 0:
                db.session.commit()

    if apprentice_sheet is not None:
        for index, data in enumerate(parse_apprentice_sheet(apprentice_sheet), start=1):
            group_number = data.get("group_number")
            group_obj = existing_groups.get(group_number)
            if not group_obj:
                group_obj = TrainingGroup(
                    group_number=group_number,
                    program_name=data.get("program_name", "SIN DEFINIR"),
                    program_level=data.get("program_level", ""),
                    lead_instructor=data.get("lead_instructor", ""),
                    followup_instructor=data.get("followup_instructor", ""),
                    created_by=owner_id
                )
                db.session.add(group_obj)
                db.session.flush()
                existing_groups[group_number] = group_obj
                result.created_groups += 1
                result.group_count += 1
            clean_data = {key: value for key, value in data.items() if key in APPRENTICE_MODEL_FIELDS and key != "group_id"}
            clean_data["group_id"] = group_obj.id
            clean_data["group_number"] = group_obj.group_number
            for inherited in ("program_name", "program_level", "lead_instructor", "followup_instructor", "group_validity"):
                if not clean_data.get(inherited):
                    clean_data[inherited] = getattr(group_obj, inherited, "") or ""
            apprentice_obj = existing_apprentices.get(data["document_number"])
            if apprentice_obj is None:
                apprentice_obj = Apprentice(created_by=owner_id, **clean_data)
                db.session.add(apprentice_obj)
                db.session.flush()
                existing_apprentices[data["document_number"]] = apprentice_obj
                result.created_apprentices += 1
            else:
                for key, value in clean_data.items():
                    if hasattr(apprentice_obj, key):
                        setattr(apprentice_obj, key, value)
                apprentice_obj.created_by = owner_id
                result.updated_apprentices += 1
            upsert_student_user(apprentice_obj, known_users=known_users)
            ensure_submissions_for_apprentice(apprentice_obj)
            result.apprentice_count += 1
            if index % 100 == 0:
                db.session.commit()

    db.session.commit()
    return result
