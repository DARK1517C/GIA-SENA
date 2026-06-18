# services/excel_export.py
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.group import TrainingGroup  # importa la constante RECORD_FIELDS desde el modelo


def format_value(value):
    """Formateador simple: fechas a YYYY-MM-DD si tienen strftime, None -> ''."""
    try:
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        return str(value)
    except Exception:
        return str(value)


def export_workbook(title, fields, rows):
    """
    Exporta una lista de objetos (rows) a un workbook con headers definidos en fields.
    fields: lista de tuplas (key, label)
    rows: lista de objetos ORM o dicts
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    headers = [label for _key, label in fields]
    sheet.append(headers)

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B8F47")

    for row in rows:
        row_values = []
        for key, _label in fields:
            if isinstance(row, dict):
                row_values.append(format_value(row.get(key, "") or ""))
            else:
                row_values.append(format_value(getattr(row, key, "") or ""))
        sheet.append(row_values)

    # Ajuste seguro de ancho de columnas (compatible con celdas fusionadas)
    for col_idx, column in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column:
            try:
                value = cell.value
            except Exception:
                value = None
            text = "" if value is None else str(value)
            if len(text) > max_length:
                max_length = len(text)
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = min(max_length + 4, 35)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def write_template_headers(sheet, top_headers, sub_headers=None):
    green_fill = PatternFill("solid", fgColor="0B8F47")
    white_font = Font(bold=True, color="FFFFFF")
    for index, value in enumerate(top_headers, start=1):
        cell = sheet.cell(1, index)
        cell.value = value
        cell.fill = green_fill
        cell.font = white_font
    if sub_headers:
        for index, value in enumerate(sub_headers, start=1):
            cell = sheet.cell(2, index)
            cell.value = value
            cell.fill = green_fill
            cell.font = white_font


def split_moments(value):
    parts = [part.strip() for part in (value or "").split("|") if part.strip()]
    while len(parts) < 4:
        parts.append("")
    return parts[:4]


def export_reference_workbook(apprentice_rows, group_rows):
    """
    Genera un workbook con dos hojas: 'Aprendices' y 'Record de fichas'.
    apprentice_rows y group_rows pueden ser listas de objetos ORM o dicts.
    """
    workbook = Workbook()
    apprentice_sheet = workbook.active
    apprentice_sheet.title = "Aprendices"

    APPRENTICE_TEMPLATE_HEADERS = [
        "N° DE FICHA",
        "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
        "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)",
        "NOMBRE DEL PROGRAMA DE FORMACIÓN",
        "NIVEL DEL PROGRAMA",
        "TIPO DE DOCUMENTO (CC, TI, CE)",
        "N° DE DOCUMENTO DEL APRENDIZ",
        "NOMBRES DEL APRENDIZ",
        "APELLIDOS DEL APRENDIZ",
        "GÉNERO",
        "TELÉFONO DEL APRENDIZ",
        "MUNICIPIO DE ORIGEN",
        "CORREO ELECTRÓNICO DEL APRENDIZ",
        "MODALIDAD ETAPA PRODUCTIVA",
        "FECHA INICIO DE PRÁCTICAS",
        "FECHA FINAL DE PRÁCTICAS",
        "MOMENTOS - SEGUIMIENTO Y/O EVALUACIÓN",
        None, None, None,
        "NOMBRE DE LA EMPRESA/ORG/INST",
        "DIRECCIÓN DE LA EMPRESA",
        "MUNICIPIO",
        "NOMBRE COFORMADOR",
        "CORREO ELECTRÓNICO DEL COFORMADOR",
        "TELÉFONO DEL COFORMADOR",
        "GESTIÓN INDIVIDUAL DEL APRENDIZ EN EP",
        "ESTADO DEL APRENDIZ EN SOFÍAPLUS",
        "RESPONSABLE DE AFILIACIÓN ARL",
        "FECHA EMISIÓN DE JUICIO EVALUATIVO EN SOFIA PLUS",
        "JUICIOS DE INGLÉS APROBADOS SI/NO",
    ]

    GROUP_TEMPLATE_SUB_HEADERS = [
        None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
        "CONTRATO DE APRENDIZAJE", "PASANTIA", "PROYECTO PRODUCTIVO", "VINCULACION LABORAL", None, None,
    ]

    write_template_headers(apprentice_sheet, APPRENTICE_TEMPLATE_HEADERS)
    try:
        apprentice_sheet.merge_cells("Q1:T1")
    except Exception:
        pass

    for apprentice in apprentice_rows:
        moments = split_moments(getattr(apprentice, "followup_moments", "") if not isinstance(apprentice, dict) else apprentice.get("followup_moments", ""))
        row = [
            getattr(apprentice, "group_number", "") if not isinstance(apprentice, dict) else apprentice.get("group_number", ""),
            getattr(apprentice, "lead_instructor", "") if not isinstance(apprentice, dict) else apprentice.get("lead_instructor", ""),
            getattr(apprentice, "followup_instructor", "") if not isinstance(apprentice, dict) else apprentice.get("followup_instructor", ""),
            getattr(apprentice, "program_name", "") if not isinstance(apprentice, dict) else apprentice.get("program_name", ""),
            getattr(apprentice, "program_level", "") if not isinstance(apprentice, dict) else apprentice.get("program_level", ""),
            getattr(apprentice, "document_type", "") if not isinstance(apprentice, dict) else apprentice.get("document_type", ""),
            getattr(apprentice, "document_number", "") if not isinstance(apprentice, dict) else apprentice.get("document_number", ""),
            getattr(apprentice, "first_names", "") if not isinstance(apprentice, dict) else apprentice.get("first_names", ""),
            getattr(apprentice, "last_names", "") if not isinstance(apprentice, dict) else apprentice.get("last_names", ""),
            getattr(apprentice, "gender", "") if not isinstance(apprentice, dict) else apprentice.get("gender", ""),
            getattr(apprentice, "phone", "") if not isinstance(apprentice, dict) else apprentice.get("phone", ""),
            getattr(apprentice, "municipality_origin", "") if not isinstance(apprentice, dict) else apprentice.get("municipality_origin", ""),
            getattr(apprentice, "email", "") if not isinstance(apprentice, dict) else apprentice.get("email", ""),
            getattr(apprentice, "ep_modality", "") if not isinstance(apprentice, dict) else apprentice.get("ep_modality", ""),
            getattr(apprentice, "practice_start_date", "") if not isinstance(apprentice, dict) else apprentice.get("practice_start_date", ""),
            getattr(apprentice, "practice_end_date", "") if not isinstance(apprentice, dict) else apprentice.get("practice_end_date", ""),
            moments[0], moments[1], moments[2], moments[3],
            getattr(apprentice, "company_name", "") if not isinstance(apprentice, dict) else apprentice.get("company_name", ""),
            getattr(apprentice, "company_address", "") if not isinstance(apprentice, dict) else apprentice.get("company_address", ""),
            getattr(apprentice, "company_municipality", "") if not isinstance(apprentice, dict) else apprentice.get("company_municipality", ""),
            getattr(apprentice, "coformador_name", "") if not isinstance(apprentice, dict) else apprentice.get("coformador_name", ""),
            getattr(apprentice, "coformador_email", "") if not isinstance(apprentice, dict) else apprentice.get("coformador_email", ""),
            getattr(apprentice, "coformador_phone", "") if not isinstance(apprentice, dict) else apprentice.get("coformador_phone", ""),
            getattr(apprentice, "individual_management", "") if not isinstance(apprentice, dict) else apprentice.get("individual_management", ""),
            getattr(apprentice, "sofia_status", "") if not isinstance(apprentice, dict) else apprentice.get("sofia_status", ""),
            getattr(apprentice, "arl_responsible", "") if not isinstance(apprentice, dict) else apprentice.get("arl_responsible", ""),
            getattr(apprentice, "evaluation_date", "") if not isinstance(apprentice, dict) else apprentice.get("evaluation_date", ""),
            getattr(apprentice, "english_results", "") if not isinstance(apprentice, dict) else apprentice.get("english_results", ""),
        ]
        apprentice_sheet.append([format_value(v) for v in row])

    # HOJA: Record de fichas (usar la constante del modelo como fuente de verdad)
    group_sheet = workbook.create_sheet("Record de fichas")
    group_headers = [label for key, label in TrainingGroup.RECORD_FIELDS]
    write_template_headers(group_sheet, group_headers, GROUP_TEMPLATE_SUB_HEADERS)

    # aplicar merges si se requieren (mantener compatibilidad con el formato anterior)
    merged_ranges = ["A1:A2", "B1:B2", "C1:C2", "D1:D2", "E1:E2", "F1:F2", "G1:G2", "H1:H2", "I1:I2", "J1:J2", "K1:K2", "L1:L2", "M1:M2", "N1:N2", "O1:O2", "P1:P2", "U1:U2", "V1:V2"]
    for merged in merged_ranges:
        try:
            group_sheet.merge_cells(merged)
        except Exception:
            pass
    try:
        group_sheet.merge_cells("Q1:T1")
    except Exception:
        pass

    for index, group in enumerate(group_rows, start=1):
        row = []
        for key, _label in TrainingGroup.RECORD_FIELDS:
            if key == "consecutive":
                row.append(index)
                continue
            if isinstance(group, dict):
                row.append(group.get(key, ""))
            else:
                row.append(getattr(group, key, ""))
        group_sheet.append([format_value(v) for v in row])

    # Ajuste seguro de ancho de columnas para todas las hojas
    for sheet in workbook.worksheets:
        for col_idx, column in enumerate(sheet.columns, start=1):
            max_length = 0
            for cell in column:
                try:
                    value = cell.value
                except Exception:
                    value = None
                text = "" if value is None else str(value)
                if len(text) > max_length:
                    max_length = len(text)
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = min(max_length + 4, 35)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
