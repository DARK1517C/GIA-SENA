import os
import re
import sqlite3
import unicodedata
from datetime import datetime
from functools import wraps
from io import BytesIO

from flask import (
    Flask,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import event, inspect, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.engine import Engine
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "gia-sena-secret")
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv(
    "DATABASE_URL",
    os.getenv("MYSQL_URL", f"sqlite:///{os.path.join(BASE_DIR, 'gia.db')}"),
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True}

if app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite"):
    app.config["SQLALCHEMY_ENGINE_OPTIONS"]["connect_args"] = {
        "timeout": 120,
        "check_same_thread": False,
    }

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"

os.makedirs(UPLOAD_DIR, exist_ok=True)


@event.listens_for(Engine, "connect")
def configure_sqlite(connection, _record):
    if isinstance(connection, sqlite3.Connection):
        cursor = connection.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA synchronous=NORMAL")
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.execute("PRAGMA temp_store=MEMORY")
        cursor.execute("PRAGMA busy_timeout=120000")
        cursor.close()


APPRENTICE_FIELDS = [
    ("group_number", "NÚMERO DE FICHA"),
    ("document_type", "TIPO DE DOCUMENTO"),
    ("document_number", "NÚMERO DE DOCUMENTO"),
    ("first_names", "NOMBRES"),
    ("last_names", "APELLIDOS"),
    ("gender", "GÉNERO"),
    ("phone", "TELÉFONO"),
    ("email", "CORREO ELECTRÓNICO"),
    ("municipality_origin", "MUNICIPIO DE ORIGEN"),
    ("program_name", "NOMBRE DEL PROGRAMA"),
    ("group_validity", "VIGENCIA FICHA"),
    ("lead_instructor", "INSTRUCTOR LÍDER"),
    ("followup_instructor", "INSTRUCTOR SEGUIMIENTO"),
    ("ep_modality", "MODALIDAD EP"),
    ("sofia_status", "ESTADO SOFIA PLUS"),
    ("practice_start_date", "FECHA INICIO PRÁCTICAS"),
    ("practice_end_date", "FECHA FINAL PRÁCTICAS"),
    ("company_name", "NOMBRE EMPRESA"),
    ("company_municipality", "MUNICIPIO EMPRESA"),
    ("company_address", "DIRECCIÓN EMPRESA"),
    ("coformador_name", "NOMBRE COFORMADOR"),
    ("coformador_email", "CORREO COFORMADOR"),
    ("coformador_phone", "TELÉFONO COFORMADOR"),
    ("arl_responsible", "RESPONSABLE ARL"),
    ("individual_management", "GESTIÓN INDIVIDUAL"),
    ("followup_moments", "MOMENTOS DE SEGUIMIENTO"),
    ("evaluation_date", "FECHA JUICIO EVALUATIVO"),
    ("english_results", "JUICIOS DE INGLÉS"),
]

GROUP_FIELDS = [
    ("group_number", "NÚMERO DE FICHA"),
    ("program_name", "NOMBRE DEL PROGRAMA"),
    ("followup_instructor", "INSTRUCTOR DE SEGUIMIENTO"),
    ("municipality", "MUNICIPIO"),
    ("program_level", "NIVEL DEL PROGRAMA"),
    ("modality", "MODALIDAD"),
    ("sofia_group_status", "ESTADO FICHA SOFIA PLUS"),
    ("group_validity", "VIGENCIA FICHA"),
    ("group_start_date", "FECHA INICIO FICHA"),
    ("training_end_date", "FECHA FIN FORMACIÓN"),
    ("ep_start_date", "FECHA INICIO EP"),
    ("apprentices_statistics", "ESTADÍSTICAS DE APRENDICES"),
    ("apprentices_training", "APRENDICES EN FORMACIÓN"),
    ("apprentices_enabled", "APRENDICES HABILITADOS"),
    ("apprentices_rap_pending", "APRENDICES DEBEN RAP"),
    ("apprentices_practice", "APRENDICES EN PRÁCTICA"),
    ("apprentices_without_alternative", "APRENDICES SIN ALTERNATIVA"),
    ("apprentices_certified", "APRENDICES CERTIFICADOS"),
    ("productive_modalities", "MODALIDADES DE ETAPA PRODUCTIVA"),
    ("learning_contract", "CONTRATO APRENDIZAJE"),
    ("internship", "PASANTÍA"),
    ("productive_project", "PROYECTO PRODUCTIVO"),
    ("employment_link", "VINCULACIÓN LABORAL"),
]

APPRENTICE_TEMPLATE_HEADERS = [
    "N° DE FICHA",
    "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
    "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)",
    "NOMBRE DEL PROGRAMA DE FORMACIÓN",
    "NIVEL DEL PROGRAMA",
    "TIPO DE DOCUMENTO (CC, TI, CE)",
    "N° DE DOCUMENTO DEL APRENDIZ",
    "NOMBRES DEL APRENDIZ",
    "APELLIDOS DEL APRENDIZ",
    "GÉNERO (F/M)",
    "TELÉFONO DEL APRENDIZ",
    "MUNICIPIO DE ORIGEN",
    "CORREO ELECTRÓNICO DEL APRENDIZ",
    "MODALIDAD ETAPA PRODUCTIVA",
    "FECHA INICIO DE PRÁCTICAS",
    "FECHA FINAL DE PRÁCTICAS",
    "MOMENTOS - SEGUIMIENTO Y/O EVALUACIÓN",
    None,
    None,
    None,
    "NOMBRE DE LA EMPRESA/ORG/INST",
    "DIRECCIÓN DE LA EMPRESA",
    "MUNICIPIO",
    "NOMBRE COFORMADOR",
    "CORREO ELECTRÓNICO DEL COFORMADOR",
    "TELÉFONO DEL COFORMADOR",
    "GESTIÓN INDIVIDUAL DEL APRENDIZ EN EP",
    "ESTADO DEL APRENDIZ EN SOFÍAPLUS",
    "RESPONSABLE DE AFILIACIÓN ARL",
    "FECHA EMISIÓN DE JUICIO EVALUATIVO EN SOFIA PLUS",
    "JUICIOS DE INGLÉS APROBADOS SI/NO",
]

GROUP_TEMPLATE_TOP_HEADERS = [
    "CONSECUTIVO",
    "N° DE FICHA",
    "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
    "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)",
    "NOMBRE DEL PROGRAMA DE FORMACIÓN",
    "MUNICIPIO",
    "NIVEL DE PROGRAMA",
    "MODALIDAD",
    "ESTADO DE LA FICHA EN SOFÍAPLUS",
    "FECHA INICIO DE LA FICHA EN SOFIAPLUS",
    "FECHA FIN DE LA FORMACIÓN EN SOFIAPLUS",
    "FECHA INICIO DE ETAPA PRODUCTIVA",
    "VIGENCIA DE LA FICHA",
    "APRENDICES EN FORMACIÓN",
    "APRENDICES HABILITADOS PARA INICIAR ETAPA PRODUCTIVA",
    "APRENDICES QUE DEBEN RAP",
    "APRENDICES EN PRÁCTICA",
    None,
    None,
    None,
    "APRENDICES SIN ALTERNATIVA DE PRÁCTIVA",
    "APRENDICES CERTIFICADOS",
]

GROUP_TEMPLATE_SUB_HEADERS = [
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    "CONTRATO DE APRENDIZAJE",
    "PASANTIA",
    "PROYECTO PRODUCTIVO",
    "VINCULACION LABORAL",
    None,
    None,
]

APPRENTICE_IMPORT_ALIASES = {
    "group_number": [
        "N° DE FICHA",
        "N DE FICHA",
        "NUMERO DE FICHA",
    ],
    "lead_instructor": [
        "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
        "NOMBRE DE INSTRUCTOR(A) LIDER DE LA FICHA",
        "INSTRUCTOR LIDER",
    ],
    "followup_instructor": [
        "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)",
        "INSTRUCTOR SEGUIMIENTO",
    ],
    "program_name": ["NOMBRE DEL PROGRAMA DE FORMACIÓN", "NOMBRE DEL PROGRAMA"],
    "document_type": ["TIPO DE DOCUMENTO (CC, TI, CE)", "TIPO DE DOCUMENTO"],
    "document_number": ["N° DE DOCUMENTO DEL APRENDIZ", "N DE DOCUMENTO DEL APRENDIZ", "NÚMERO DE DOCUMENTO"],
    "first_names": ["NOMBRES DEL APRENDIZ", "NOMBRES"],
    "last_names": ["APELLIDOS DEL APRENDIZ", "APELLIDOS"],
    "gender": ["GÉNERO (F/M)", "GENERO (F/M)", "GÉNERO"],
    "phone": ["TELÉFONO DEL APRENDIZ", "TELEFONO DEL APRENDIZ", "TELÉFONO"],
    "municipality_origin": ["MUNICIPIO DE ORIGEN"],
    "email": ["CORREO ELECTRÓNICO DEL APRENDIZ", "CORREO ELECTRONICO DEL APRENDIZ", "CORREO ELECTRÓNICO"],
    "ep_modality": ["MODALIDAD ETAPA PRODUCTIVA", "MODALIDAD EP"],
    "practice_start_date": ["FECHA INICIO DE PRÁCTICAS"],
    "practice_end_date": ["FECHA FINAL DE PRÁCTICAS"],
    "company_name": ["NOMBRE DE LA EMPRESA/ORG/INST", "NOMBRE EMPRESA"],
    "company_address": ["DIRECCIÓN DE LA EMPRESA", "DIRECCION DE LA EMPRESA", "DIRECCIÓN EMPRESA"],
    "company_municipality": ["MUNICIPIO", "MUNICIPIO EMPRESA"],
    "coformador_name": ["NOMBRE COFORMADOR"],
    "coformador_email": ["CORREO ELECTRÓNICO DEL COFORMADOR", "CORREO ELECTRONICO DEL COFORMADOR", "CORREO COFORMADOR"],
    "coformador_phone": ["TELÉFONO DEL COFORMADOR", "TELEFONO DEL COFORMADOR", "TELÉFONO COFORMADOR"],
    "individual_management": ["GESTIÓN INDIVIDUAL DEL APRENDIZ EN EP", "GESTION INDIVIDUAL DEL APRENDIZ EN EP", "GESTIÓN INDIVIDUAL"],
    "sofia_status": ["ESTADO DEL APRENDIZ EN SOFÍAPLUS", "ESTADO DEL APRENDIZ EN SOFIAPLUS", "ESTADO SOFIA PLUS"],
    "arl_responsible": ["RESPONSABLE DE AFILIACIÓN ARL", "RESPONSABLE DE AFILIACION ARL", "RESPONSABLE ARL"],
    "evaluation_date": [
        "FECHA EMISIÓN DE JUICIO EVALUATIVO EN SOFIA PLUS",
        "FECHA EMISION DE JUICIO EVALUATIVO EN SOFIA PLUS",
        "FECHA JUICIO EVALUATIVO",
    ],
    "english_results": ["JUICIOS DE INGLÉS APROBADOS SI/NO", "JUICIOS DE INGLES APROBADOS SI/NO", "JUICIOS DE INGLÉS"],
    "program_level": ["NIVEL DEL PROGRAMA", "NIVEL DE PROGRAMA"],
}

GROUP_IMPORT_ALIASES = {
    "group_number": ["N° DE FICHA", "N DE FICHA"],
    "lead_instructor": [
        "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
        "NOMBRE DE INSTRUCTOR(A) LIDER DE LA FICHA",
    ],
    "followup_instructor": ["NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)"],
    "program_name": ["NOMBRE DEL PROGRAMA DE FORMACIÓN"],
    "municipality": ["MUNICIPIO"],
    "program_level": ["NIVEL DE PROGRAMA"],
    "modality": ["MODALIDAD"],
    "sofia_group_status": ["ESTADO DE LA FICHA EN SOFÍAPLUS", "ESTADO DE LA FICHA EN SOFIAPLUS"],
    "group_start_date": ["FECHA INICIO DE LA FICHA EN SOFIAPLUS"],
    "training_end_date": ["FECHA FIN DE LA FORMACIÓN EN SOFIAPLUS", "FECHA FIN DE LA FORMACION EN SOFIAPLUS"],
    "ep_start_date": ["FECHA INICIO DE ETAPA PRODUCTIVA"],
    "group_validity": ["VIGENCIA DE LA FICHA"],
    "apprentices_training": ["APRENDICES EN FORMACIÓN", "APRENDICES EN FORMACION"],
    "apprentices_enabled": ["APRENDICES HABILITADOS PARA INICIAR ETAPA PRODUCTIVA", "APRENDICES HABILITADOS"],
    "apprentices_rap_pending": ["APRENDICES QUE DEBEN RAP", "APRENDICES DEBEN RAP"],
    "apprentices_practice": ["APRENDICES EN PRÁCTICA", "APRENDICES EN PRACTICA"],
    "apprentices_without_alternative": ["APRENDICES SIN ALTERNATIVA DE PRÁCTIVA", "APRENDICES SIN ALTERNATIVA DE PRACTIVA"],
    "apprentices_certified": ["APRENDICES CERTIFICADOS"],
    "learning_contract": ["CONTRATO DE APRENDIZAJE", "CONTRATO APRENDIZAJE"],
    "internship": ["PASANTIA", "PASANTÍA"],
    "productive_project": ["PROYECTO PRODUCTIVO"],
    "employment_link": ["VINCULACION LABORAL", "VINCULACIÓN LABORAL"],
}

USER_ROLES = {
    "docente": "Docente",
    "visualizador": "Visualizador",
    "super_admin": "Super Admin",
    "aprendiz": "Aprendiz",
}

APPRENTICE_MODEL_FIELDS = [key for key, _label in APPRENTICE_FIELDS]
GROUP_MODEL_FIELDS = [
    "group_number",
    "program_name",
    "lead_instructor",
    "followup_instructor",
    "municipality",
    "program_level",
    "modality",
    "sofia_group_status",
    "group_validity",
    "group_start_date",
    "training_end_date",
    "ep_start_date",
    "apprentices_statistics",
    "apprentices_training",
    "apprentices_enabled",
    "apprentices_rap_pending",
    "apprentices_practice",
    "apprentices_without_alternative",
    "apprentices_certified",
    "productive_modalities",
    "learning_contract",
    "internship",
    "productive_project",
    "employment_link",
]


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(30), nullable=False, default="docente")
    full_name = db.Column(db.String(150), nullable=False)
    email = db.Column(db.String(150), nullable=True)
    document_type = db.Column(db.String(20), nullable=True)
    document_number = db.Column(db.String(30), nullable=True)
    managed_group_numbers = db.Column(db.Text, nullable=True)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    apprentices = db.relationship(
        "Apprentice",
        foreign_keys="Apprentice.created_by",
        backref="owner",
        lazy=True,
    )
    groups = db.relationship(
        "TrainingGroup",
        foreign_keys="TrainingGroup.created_by",
        backref="creator",
        lazy=True,
    )

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class TrainingGroup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)

    group_number = db.Column(db.String(30), unique=True, nullable=False)
    program_name = db.Column(db.String(150), nullable=False)
    lead_instructor = db.Column(db.String(150), nullable=True)
    followup_instructor = db.Column(db.String(150), nullable=True)
    municipality = db.Column(db.String(120), nullable=True)
    program_level = db.Column(db.String(80), nullable=True)
    modality = db.Column(db.String(80), nullable=True)
    sofia_group_status = db.Column(db.String(80), nullable=True)
    group_validity = db.Column(db.String(80), nullable=True)
    group_start_date = db.Column(db.String(40), nullable=True)
    training_end_date = db.Column(db.String(40), nullable=True)
    ep_start_date = db.Column(db.String(40), nullable=True)
    apprentices_statistics = db.Column(db.String(120), nullable=True)
    apprentices_training = db.Column(db.String(30), nullable=True)
    apprentices_enabled = db.Column(db.String(30), nullable=True)
    apprentices_rap_pending = db.Column(db.String(30), nullable=True)
    apprentices_practice = db.Column(db.String(30), nullable=True)
    apprentices_without_alternative = db.Column(db.String(30), nullable=True)
    apprentices_certified = db.Column(db.String(30), nullable=True)
    productive_modalities = db.Column(db.String(120), nullable=True)
    learning_contract = db.Column(db.String(30), nullable=True)
    internship = db.Column(db.String(30), nullable=True)
    productive_project = db.Column(db.String(30), nullable=True)
    employment_link = db.Column(db.String(30), nullable=True)


class Apprentice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    student_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)

    group_number = db.Column(db.String(30), nullable=False)
    document_type = db.Column(db.String(30), nullable=False)
    document_number = db.Column(db.String(30), unique=True, nullable=False)
    first_names = db.Column(db.String(120), nullable=False)
    last_names = db.Column(db.String(120), nullable=False)
    gender = db.Column(db.String(20), nullable=True)
    phone = db.Column(db.String(30), nullable=True)
    email = db.Column(db.String(150), nullable=True)
    municipality_origin = db.Column(db.String(120), nullable=True)
    program_name = db.Column(db.String(150), nullable=True)
    group_validity = db.Column(db.String(80), nullable=True)
    lead_instructor = db.Column(db.String(150), nullable=True)
    followup_instructor = db.Column(db.String(150), nullable=True)
    ep_modality = db.Column(db.String(120), nullable=True)
    sofia_status = db.Column(db.String(80), nullable=True)
    practice_start_date = db.Column(db.String(40), nullable=True)
    practice_end_date = db.Column(db.String(40), nullable=True)
    company_name = db.Column(db.String(150), nullable=True)
    company_municipality = db.Column(db.String(120), nullable=True)
    company_address = db.Column(db.String(180), nullable=True)
    coformador_name = db.Column(db.String(150), nullable=True)
    coformador_email = db.Column(db.String(150), nullable=True)
    coformador_phone = db.Column(db.String(30), nullable=True)
    arl_responsible = db.Column(db.String(150), nullable=True)
    individual_management = db.Column(db.Text, nullable=True)
    followup_moments = db.Column(db.String(120), nullable=True)
    evaluation_date = db.Column(db.String(40), nullable=True)
    english_results = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    bitacoras = db.relationship(
        "Bitacora",
        backref="apprentice",
        lazy=True,
        cascade="all, delete-orphan",
        order_by="desc(Bitacora.created_at)",
    )
    student_user = db.relationship("User", foreign_keys=[student_user_id], lazy=True)

    @property
    def full_name(self):
        return f"{self.first_names} {self.last_names}".strip()


class Bitacora(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    apprentice_id = db.Column(db.Integer, db.ForeignKey("apprentice.id"), nullable=False)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    title = db.Column(db.String(120), nullable=False)
    notes = db.Column(db.Text, nullable=True)
    file_name = db.Column(db.String(255), nullable=True)
    file_path = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    uploaded_by = db.relationship("User", lazy=True)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def role_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role not in roles:
                abort(403)
            return view(*args, **kwargs)

        return wrapped

    return decorator


def parse_form(fields):
    data = {}
    for key, _label in fields:
        data[key] = request.form.get(key, "").strip()
    return data


def normalize_header(value):
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ").strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"\s+", " ", text)
    return text


def clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def build_alias_lookup(mapping):
    lookup = {}
    for key, aliases in mapping.items():
        for alias in aliases:
            lookup[normalize_header(alias)] = key
    return lookup


APPRENTICE_ALIAS_LOOKUP = build_alias_lookup(APPRENTICE_IMPORT_ALIASES)
GROUP_ALIAS_LOOKUP = build_alias_lookup(GROUP_IMPORT_ALIASES)


def export_workbook(title, fields, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title

    headers = [label for _key, label in fields]
    sheet.append(headers)
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B8F47")

    for row in rows:
        sheet.append([getattr(row, key, "") or "" for key, _label in fields])

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 35)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def find_sheet_by_headers(workbook, required_headers, alias_lookup, min_matches=4):
    best_match = None
    best_score = 0
    required = {normalize_header(item) for item in required_headers}
    for sheet in workbook.worksheets:
        for row_index in range(1, min(sheet.max_row, 8) + 1):
            values = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
            normalized = [normalize_header(value) for value in values if normalize_header(value)]
            matches = sum(1 for item in normalized if item in required or item in alias_lookup)
            if matches > best_score:
                best_score = matches
                best_match = (sheet, row_index)
    if best_match and best_score >= min_matches:
        return best_match
    return None, None


def extract_sheet_rows(sheet, header_row):
    headers = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
    data = []
    for row_index in range(header_row + 1, sheet.max_row + 1):
        row_values = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "", "\xa0") for value in row_values):
            continue
        data.append(row_values)
    return headers, data


def parse_apprentice_sheet(sheet, header_row):
    headers, data_rows = extract_sheet_rows(sheet, header_row)
    header_map = {index + 1: normalize_header(value) for index, value in enumerate(headers) if normalize_header(value)}
    records = {}

    for row in data_rows:
        record = {key: "" for key in APPRENTICE_MODEL_FIELDS}
        for col_index, normalized in header_map.items():
            field = APPRENTICE_ALIAS_LOOKUP.get(normalized)
            if not field:
                continue
            value = clean_cell(row[col_index - 1] if col_index - 1 < len(row) else "")
            if field in record:
                record[field] = value

        moments = []
        for col_index in range(17, 21):
            if col_index - 1 < len(row):
                value = clean_cell(row[col_index - 1])
                if value:
                    moments.append(value)
        if moments:
            record["followup_moments"] = " | ".join(moments)

        if record["document_number"]:
            records[record["document_number"]] = record
    return list(records.values())


def parse_group_sheet(sheet, header_row):
    headers = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
    subheaders = [sheet.cell(header_row + 1, col).value for col in range(1, sheet.max_column + 1)] if sheet.max_row > header_row else []
    header_map = {}
    for index, value in enumerate(headers, start=1):
        normalized = normalize_header(value)
        if normalized:
            header_map[index] = normalized
    for index, value in enumerate(subheaders, start=1):
        normalized = normalize_header(value)
        if normalized:
            header_map[index] = normalized

    start_data_row = header_row + 2 if any(normalize_header(item) for item in subheaders) else header_row + 1
    records = {}

    for row_index in range(start_data_row, sheet.max_row + 1):
        row = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "", "\xa0") for value in row):
            continue

        record = {key: "" for key, _label in GROUP_FIELDS}
        for col_index, normalized in header_map.items():
            field = GROUP_ALIAS_LOOKUP.get(normalized)
            if not field:
                continue
            record[field] = clean_cell(row[col_index - 1] if col_index - 1 < len(row) else "")

        modality_parts = []
        for key, label in [
            ("learning_contract", "Contrato aprendizaje"),
            ("internship", "Pasantía"),
            ("productive_project", "Proyecto productivo"),
            ("employment_link", "Vinculación laboral"),
        ]:
            if record[key]:
                modality_parts.append(f"{label}: {record[key]}")
        record["productive_modalities"] = " | ".join(modality_parts)
        total_stats = [
            record["apprentices_training"],
            record["apprentices_enabled"],
            record["apprentices_rap_pending"],
            record["apprentices_practice"],
            record["apprentices_without_alternative"],
            record["apprentices_certified"],
        ]
        record["apprentices_statistics"] = " / ".join(value for value in total_stats if value)

        if record["group_number"]:
            records[record["group_number"]] = record
    return list(records.values())


def import_reference_workbook(file_storage, owner_id, mode="both"):
    workbook = load_workbook(file_storage, data_only=False)
    apprentice_sheet = apprentice_header = None
    group_sheet = group_header = None

    if mode in {"both", "apprentices"}:
        apprentice_sheet, apprentice_header = find_sheet_by_headers(
            workbook,
            [
                "N° DE DOCUMENTO DEL APRENDIZ",
                "NOMBRES DEL APRENDIZ",
                "MODALIDAD ETAPA PRODUCTIVA",
                "GESTIÓN INDIVIDUAL DEL APRENDIZ EN EP",
            ],
            APPRENTICE_ALIAS_LOOKUP,
        )
    if mode in {"both", "groups"}:
        group_sheet, group_header = find_sheet_by_headers(
            workbook,
            [
                "N° DE FICHA",
                "APRENDICES EN FORMACIÓN",
                "APRENDICES EN PRÁCTICA",
                "APRENDICES CERTIFICADOS",
            ],
            GROUP_ALIAS_LOOKUP,
        )

    apprentice_count = 0
    group_count = 0
    existing_apprentices = {item.document_number: item for item in Apprentice.query.all()}
    existing_groups = {item.group_number: item for item in TrainingGroup.query.all()}
    known_users = {item.username: item for item in User.query.filter_by(role="aprendiz").all()}

    if apprentice_sheet is not None:
        for index, data in enumerate(parse_apprentice_sheet(apprentice_sheet, apprentice_header), start=1):
            clean_data = {key: value for key, value in data.items() if key in APPRENTICE_MODEL_FIELDS}
            apprentice = existing_apprentices.get(data["document_number"])
            if apprentice is None:
                apprentice = Apprentice(created_by=owner_id, **clean_data)
                db.session.add(apprentice)
                existing_apprentices[data["document_number"]] = apprentice
            else:
                for key, value in clean_data.items():
                    setattr(apprentice, key, value)
                apprentice.created_by = owner_id
            upsert_student_user(apprentice, known_users=known_users)
            apprentice_count += 1
            if index % 100 == 0:
                db.session.commit()

    if group_sheet is not None:
        for index, data in enumerate(parse_group_sheet(group_sheet, group_header), start=1):
            clean_data = {key: value for key, value in data.items() if key in GROUP_MODEL_FIELDS}
            group = existing_groups.get(data["group_number"])
            if group is None:
                group = TrainingGroup(created_by=owner_id, **clean_data)
                db.session.add(group)
                existing_groups[data["group_number"]] = group
            else:
                for key, value in clean_data.items():
                    if hasattr(group, key):
                        setattr(group, key, value)
                group.created_by = owner_id
            group_count += 1
            if index % 100 == 0:
                db.session.commit()

    db.session.commit()
    return apprentice_count, group_count, apprentice_sheet is not None, group_sheet is not None


def split_moments(value):
    parts = [part.strip() for part in (value or "").split("|") if part.strip()]
    while len(parts) < 4:
        parts.append("")
    return parts[:4]


def write_template_headers(sheet, top_headers, sub_headers=None):
    green_fill = PatternFill("solid", fgColor="0B8F47")
    white_font = Font(bold=True, color="FFFFFF")

    for index, value in enumerate(top_headers, start=1):
        cell = sheet.cell(1, index)
        cell.value = value
        cell.fill = green_fill
        cell.font = white_font

    if sub_headers:
        for index, value in enumerate(sub_headers, start=1):
            cell = sheet.cell(2, index)
            cell.value = value
            cell.fill = green_fill
            cell.font = white_font


def export_reference_workbook(apprentice_rows, group_rows):
    workbook = Workbook()
    apprentice_sheet = workbook.active
    apprentice_sheet.title = "Aprendices"
    group_sheet = workbook.create_sheet("Record de fichas")

    write_template_headers(apprentice_sheet, APPRENTICE_TEMPLATE_HEADERS)
    apprentice_sheet.merge_cells("Q1:T1")

    for row_index, apprentice in enumerate(apprentice_rows, start=2):
        moments = split_moments(apprentice.followup_moments)
        apprentice_sheet.append(
            [
                apprentice.group_number,
                apprentice.lead_instructor,
                apprentice.followup_instructor,
                apprentice.program_name,
                "",
                apprentice.document_type,
                apprentice.document_number,
                apprentice.first_names,
                apprentice.last_names,
                apprentice.gender,
                apprentice.phone,
                apprentice.municipality_origin,
                apprentice.email,
                apprentice.ep_modality,
                apprentice.practice_start_date,
                apprentice.practice_end_date,
                moments[0],
                moments[1],
                moments[2],
                moments[3],
                apprentice.company_name,
                apprentice.company_address,
                apprentice.company_municipality,
                apprentice.coformador_name,
                apprentice.coformador_email,
                apprentice.coformador_phone,
                apprentice.individual_management,
                apprentice.sofia_status,
                apprentice.arl_responsible,
                apprentice.evaluation_date,
                apprentice.english_results,
            ]
        )

    write_template_headers(group_sheet, GROUP_TEMPLATE_TOP_HEADERS, GROUP_TEMPLATE_SUB_HEADERS)
    for merged in ["A1:A2", "B1:B2", "C1:C2", "D1:D2", "E1:E2", "F1:F2", "G1:G2", "H1:H2", "I1:I2", "J1:J2", "K1:K2", "L1:L2", "M1:M2", "N1:N2", "O1:O2", "P1:P2", "U1:U2", "V1:V2"]:
        group_sheet.merge_cells(merged)
    group_sheet.merge_cells("Q1:T1")

    for index, group in enumerate(group_rows, start=1):
        group_sheet.append(
            [
                index,
                group.group_number,
                getattr(group, "lead_instructor", ""),
                group.followup_instructor,
                group.program_name,
                group.municipality,
                group.program_level,
                group.modality,
                group.sofia_group_status,
                group.group_start_date,
                group.training_end_date,
                group.ep_start_date,
                group.group_validity,
                group.apprentices_training,
                group.apprentices_enabled,
                group.apprentices_rap_pending,
                group.apprentices_practice,
                group.learning_contract,
                group.internship,
                group.productive_project,
                group.apprentices_without_alternative,
                group.apprentices_certified,
            ]
        )
    for sheet in workbook.worksheets:
        for column_index, column in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 4, 34)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def model_from_headers(headers, fields):
    normalized = {str(header).strip().upper(): index for index, header in enumerate(headers)}
    return [normalized.get(label.upper()) for _key, label in fields]


def can_access_apprentice(apprentice):
    if current_user.role in {"super_admin", "visualizador"}:
        return True
    if current_user.role == "docente":
        managed_numbers = get_user_managed_group_numbers()
        if managed_numbers:
            return apprentice.group_number in managed_numbers
        return apprentice.created_by == current_user.id
    if current_user.role == "aprendiz":
        return apprentice.student_user_id == current_user.id
    return False


def get_apprentice_or_404(apprentice_id):
    apprentice = Apprentice.query.get_or_404(apprentice_id)
    if not can_access_apprentice(apprentice):
        abort(403)
    return apprentice


def can_access_group(group):
    if current_user.role in {"super_admin", "visualizador"}:
        return True
    if current_user.role == "docente":
        managed_numbers = get_user_managed_group_numbers()
        if managed_numbers:
            return group.group_number in managed_numbers
        return group.created_by == current_user.id
    return False


def get_group_or_404(group_id):
    group = TrainingGroup.query.get_or_404(group_id)
    if not can_access_group(group):
        abort(403)
    return group


def build_apprentice_query():
    search = request.args.get("search", "").strip()
    group_filter = request.args.get("group_number", "").strip()
    modality = request.args.get("ep_modality", "").strip()
    status = request.args.get("status", "").strip()

    query = visible_apprentices_query()

    if search:
        like = f"%{search}%"
        query = query.filter(
            (Apprentice.first_names.ilike(like))
            | (Apprentice.last_names.ilike(like))
            | (Apprentice.document_number.ilike(like))
            | (Apprentice.group_number.ilike(like))
        )
    if group_filter:
        query = query.filter(Apprentice.group_number.ilike(f"%{group_filter}%"))
    if modality:
        query = query.filter(Apprentice.ep_modality.ilike(f"%{modality}%"))
    if status:
        query = query.filter(Apprentice.sofia_status.ilike(f"%{status}%"))
    return query


def build_group_query():
    search = request.args.get("search", "").strip()
    quick_search = request.args.get("quick_search", "").strip()
    municipality = request.args.get("municipality", "").strip()
    modality = request.args.get("modality", "").strip()

    query = visible_groups_query()

    if search:
        like = f"%{search}%"
        query = query.filter(
            (TrainingGroup.group_number.ilike(like))
            | (TrainingGroup.program_name.ilike(like))
            | (TrainingGroup.lead_instructor.ilike(like))
            | (TrainingGroup.followup_instructor.ilike(like))
        )
    if quick_search:
        quick_like = f"%{quick_search}%"
        query = query.filter(
            (TrainingGroup.group_number.ilike(quick_like))
            | (TrainingGroup.program_name.ilike(quick_like))
            | (TrainingGroup.lead_instructor.ilike(quick_like))
            | (TrainingGroup.followup_instructor.ilike(quick_like))
            | (TrainingGroup.municipality.ilike(quick_like))
            | (TrainingGroup.modality.ilike(quick_like))
        )
    if municipality:
        query = query.filter(TrainingGroup.municipality.ilike(f"%{municipality}%"))
    if modality:
        query = query.filter(TrainingGroup.modality.ilike(f"%{modality}%"))
    return query


def upsert_student_user(apprentice, known_users=None):
    password_base = apprentice.document_number or "gia123"
    student_user = apprentice.student_user

    if student_user is None:
        if known_users is not None:
            student_user = known_users.get(apprentice.document_number)
        else:
            student_user = User.query.filter_by(username=apprentice.document_number).first()

    if student_user is None:
        student_user = User(
            username=apprentice.document_number,
            role="aprendiz",
            full_name=apprentice.full_name,
            email=apprentice.email,
            document_type=apprentice.document_type,
            document_number=apprentice.document_number,
        )
        student_user.set_password(password_base)
        db.session.add(student_user)
        if known_users is not None:
            known_users[apprentice.document_number] = student_user

    student_user.full_name = apprentice.full_name
    student_user.email = apprentice.email
    student_user.document_type = apprentice.document_type
    student_user.document_number = apprentice.document_number
    student_user.username = apprentice.document_number
    student_user.role = "aprendiz"
    apprentice.student_user = student_user


def seed_data():
    if User.query.filter_by(username="admin").first():
        return

    admin = User(
        username="admin",
        role="super_admin",
        full_name="Administrador GIA",
        email="admin@gia.local",
        document_type="CC",
        document_number="1000000000",
    )
    admin.set_password("admin123")

    docente = User(
        username="docente1",
        role="docente",
        full_name="Carlos Alberto Ramirez Perez",
        email="docente@gia.local",
        document_type="CC",
        document_number="1000000001",
    )
    docente.set_password("docente123")

    visualizador = User(
        username="visualizador1",
        role="visualizador",
        full_name="Maria Fernanda Lopez Garcia",
        email="visualizador@gia.local",
        document_type="CC",
        document_number="1000000002",
    )
    visualizador.set_password("visual123")

    db.session.add_all([admin, docente, visualizador])
    db.session.flush()

    group = TrainingGroup(
        created_by=docente.id,
        group_number="2558934",
        program_name="Tecnologo en Analisis y Desarrollo de Software",
        lead_instructor="German Gabriel Leal Florez",
        followup_instructor=docente.full_name,
        municipality="Cisneros",
        program_level="Tecnologo",
        modality="Presencial",
        sofia_group_status="Activa",
        group_validity="2026",
        group_start_date="2024-02-01",
        training_end_date="2026-06-30",
        ep_start_date="2026-01-20",
        apprentices_statistics="8 aprendices",
        apprentices_training="1",
        apprentices_enabled="8",
        apprentices_rap_pending="1",
        apprentices_practice="6",
        apprentices_without_alternative="2",
        apprentices_certified="1",
        productive_modalities="Contrato, Pasantia, Proyecto productivo",
        learning_contract="3",
        internship="2",
        productive_project="2",
        employment_link="1",
    )
    db.session.add(group)
    db.session.flush()

    apprentice = Apprentice(
        created_by=docente.id,
        group_number="2558934",
        document_type="CC",
        document_number="1003456789",
        first_names="Juan Carlos",
        last_names="Rodriguez Martinez",
        gender="Masculino",
        phone="3001234567",
        email="juan.rodriguez@misena.edu.co",
        municipality_origin="Cisneros",
        program_name=group.program_name,
        group_validity="2026",
        lead_instructor="German Gabriel Leal Florez",
        followup_instructor=docente.full_name,
        ep_modality="Contrato de aprendizaje",
        sofia_status="Por certificar",
        practice_start_date="2026-01-20",
        practice_end_date="2026-06-30",
        company_name="Empresa Demo SAS",
        company_municipality="Medellin",
        company_address="Cra 45 # 10 - 50",
        coformador_name="Laura Medina",
        coformador_email="laura@empresa.com",
        coformador_phone="3019876543",
        arl_responsible="SURA",
        individual_management="Seguimiento satisfactorio.",
        followup_moments="Etapa 1 completada",
        evaluation_date="2026-05-10",
        english_results="Aprobado",
    )
    db.session.add(apprentice)
    db.session.flush()
    upsert_student_user(apprentice)
    db.session.commit()


def ensure_schema_updates():
    expected_columns = {
        "user": {
            "managed_group_numbers": "TEXT",
        },
        "training_group": {
            "lead_instructor": "VARCHAR(150)",
        },
    }

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())

    for table_name, columns in expected_columns.items():
        if table_name not in existing_tables:
            continue

        existing_columns = {column["name"] for column in inspector.get_columns(table_name)}
        for column_name, column_type in columns.items():
            if column_name in existing_columns:
                continue
            db.session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}"))
        db.session.commit()


@app.context_processor
def inject_helpers():
    return {
        "role_labels": USER_ROLES,
        "is_docente": current_user.is_authenticated and current_user.role == "docente",
        "is_visualizador": current_user.is_authenticated and current_user.role == "visualizador",
        "is_admin": current_user.is_authenticated and current_user.role == "super_admin",
        "is_aprendiz": current_user.is_authenticated and current_user.role == "aprendiz",
    }


def normalize_group_numbers(raw_value):
    if not raw_value:
        return []
    values = []
    seen = set()
    for item in raw_value.replace("\r", "\n").replace(",", "\n").split("\n"):
        cleaned = item.strip()
        if not cleaned:
            continue
        if cleaned in seen:
            continue
        seen.add(cleaned)
        values.append(cleaned)
    return values


def get_user_managed_group_numbers(user=None):
    target_user = user or current_user
    if not getattr(target_user, "is_authenticated", False):
        return []
    return normalize_group_numbers(getattr(target_user, "managed_group_numbers", "") or "")


def apply_docente_group_scope(query, group_model, owner_field_name="created_by", group_field_name="group_number"):
    managed_numbers = get_user_managed_group_numbers()
    if managed_numbers:
        return query.filter(getattr(group_model, group_field_name).in_(managed_numbers))
    return query.filter(getattr(group_model, owner_field_name) == current_user.id)


def visible_apprentices_query():
    query = Apprentice.query
    if current_user.role == "docente":
        query = apply_docente_group_scope(query, Apprentice)
    elif current_user.role == "aprendiz":
        query = query.filter_by(student_user_id=current_user.id)
    return query


def visible_groups_query():
    query = TrainingGroup.query
    if current_user.role == "docente":
        query = apply_docente_group_scope(query, TrainingGroup)
    elif current_user.role == "aprendiz":
        numbers = [item.group_number for item in Apprentice.query.filter_by(student_user_id=current_user.id).all()]
        query = query.filter(TrainingGroup.group_number.in_(numbers or [""]))
    return query


def visible_bitacoras_query():
    apprentice_ids = [item.id for item in visible_apprentices_query().all()]
    if not apprentice_ids:
        return Bitacora.query.filter(Bitacora.id == -1)
    return Bitacora.query.filter(Bitacora.apprentice_id.in_(apprentice_ids))


@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        user = User.query.filter(
            (User.username == username) | (User.document_number == username)
        ).first()

        if user and user.check_password(password) and user.active:
            login_user(user)
            flash("Bienvenido a GIA.", "success")
            return redirect(url_for("dashboard"))

        flash("Credenciales inválidas.", "error")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Sesión finalizada correctamente.", "success")
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    apprentices = visible_apprentices_query().all()
    groups = visible_groups_query().all()

    stats = {
        "total_apprentices": len(apprentices),
        "total_groups": len(groups),
        "in_practice": sum(1 for item in apprentices if "pract" in (item.sofia_status or "").lower() or item.ep_modality),
        "without_alternative": sum(
            1 for item in apprentices if "sin alternativa" in (item.individual_management or "").lower()
        ),
        "certified": sum(1 for item in apprentices if "cert" in (item.sofia_status or "").lower()),
        "contracts": sum(1 for item in apprentices if "contrato" in (item.ep_modality or "").lower()),
        "internships": sum(1 for item in apprentices if "pasant" in (item.ep_modality or "").lower()),
        "projects": sum(1 for item in apprentices if "proyecto" in (item.ep_modality or "").lower()),
        "employment": sum(1 for item in apprentices if "vincul" in (item.ep_modality or "").lower()),
    }

    tracking_cards = [
        {"title": "Seguimiento etapa 1", "date": "Fecha por especificar"},
        {"title": "Seguimiento etapa 2", "date": "Fecha por especificar"},
        {"title": "Seguimiento etapa 3", "date": "Fecha por especificar"},
    ]
    recent_bitacoras = visible_bitacoras_query().order_by(Bitacora.created_at.desc()).limit(5).all()
    student_dashboard = None
    if current_user.role == "aprendiz":
        own_apprentice = apprentices[0] if apprentices else None
        own_bitacoras = own_apprentice.bitacoras if own_apprentice else []
        student_dashboard = {
            "pending": 0 if own_bitacoras else 1,
            "uploaded": len(own_bitacoras),
            "group_number": own_apprentice.group_number if own_apprentice else "Sin ficha",
            "last_upload": own_bitacoras[0].created_at if own_bitacoras else None,
        }
    return render_template(
        "dashboard.html",
        stats=stats,
        tracking_cards=tracking_cards,
        recent_bitacoras=recent_bitacoras,
        student_dashboard=student_dashboard,
    )


@app.route("/aprendices")
@login_required
def apprentices():
    if current_user.role == "aprendiz":
        return redirect(url_for("profile"))
    query = build_apprentice_query()
    records = query.order_by(Apprentice.created_at.desc()).all()
    groups = visible_groups_query().order_by(TrainingGroup.group_number.asc()).all()
    return render_template("apprentices.html", apprentices=records, groups=groups)


@app.route("/aprendices/nuevo", methods=["GET", "POST"])
@login_required
@role_required("docente", "super_admin")
def apprentice_create():
    if request.method == "POST":
        data = parse_form(APPRENTICE_FIELDS)
        apprentice = Apprentice(created_by=current_user.id, **data)
        db.session.add(apprentice)
        db.session.flush()
        upsert_student_user(apprentice)
        db.session.commit()
        flash("Aprendiz creado correctamente.", "success")
        return redirect(url_for("apprentices"))

    return render_template(
        "record_form.html",
        page_title="Nuevo aprendiz",
        submit_label="Guardar aprendiz",
        fields=APPRENTICE_FIELDS,
        record=None,
        cancel_url=url_for("apprentices"),
    )


@app.route("/aprendices/<int:apprentice_id>")
@login_required
def apprentice_detail(apprentice_id):
    apprentice = get_apprentice_or_404(apprentice_id)
    if current_user.role == "aprendiz":
        return redirect(url_for("profile"))
    return render_template(
        "detail.html",
        page_title="Detalle del aprendiz",
        record=apprentice,
        fields=APPRENTICE_FIELDS,
        edit_url=url_for("apprentice_edit", apprentice_id=apprentice.id)
        if current_user.role in {"docente", "super_admin"}
        else None,
        back_url=url_for("apprentices"),
    )


@app.route("/aprendices/<int:apprentice_id>/editar", methods=["GET", "POST"])
@login_required
def apprentice_edit(apprentice_id):
    apprentice = get_apprentice_or_404(apprentice_id)
    if current_user.role == "aprendiz":
        return redirect(url_for("profile"))
    if current_user.role not in {"docente", "super_admin"} and not (
        current_user.role == "aprendiz" and apprentice.student_user_id == current_user.id
    ):
        abort(403)

    if request.method == "POST":
        data = parse_form(APPRENTICE_FIELDS)
        if current_user.role == "aprendiz":
            allowed = {"phone", "email", "municipality_origin"}
            for key in allowed:
                setattr(apprentice, key, data[key])
        else:
            for key in data:
                setattr(apprentice, key, data[key])
        upsert_student_user(apprentice)
        db.session.commit()
        flash("Información actualizada.", "success")
        return redirect(url_for("apprentice_detail", apprentice_id=apprentice.id))

    fields = APPRENTICE_FIELDS
    if current_user.role == "aprendiz":
        fields = [
            ("phone", "TELÉFONO"),
            ("email", "CORREO ELECTRÓNICO"),
            ("municipality_origin", "MUNICIPIO DE ORIGEN"),
        ]
    return render_template(
        "record_form.html",
        page_title="Editar aprendiz",
        submit_label="Guardar cambios",
        fields=fields,
        record=apprentice,
        cancel_url=url_for("apprentice_detail", apprentice_id=apprentice.id),
    )


@app.route("/aprendices/<int:apprentice_id>/eliminar", methods=["POST"])
@login_required
@role_required("docente", "super_admin")
def apprentice_delete(apprentice_id):
    apprentice = get_apprentice_or_404(apprentice_id)
    db.session.delete(apprentice)
    db.session.commit()
    flash("Aprendiz eliminado.", "success")
    return redirect(url_for("apprentices"))


@app.route("/aprendices/eliminar-multiples", methods=["POST"])
@login_required
@role_required("docente", "super_admin")
def apprentice_bulk_delete():
    apprentice_ids = request.form.getlist("selected_ids")
    if not apprentice_ids:
        flash("Selecciona al menos un aprendiz para eliminar.", "error")
        return redirect(url_for("apprentices"))

    deleted = 0
    for apprentice_id in apprentice_ids:
        apprentice = get_apprentice_or_404(int(apprentice_id))
        db.session.delete(apprentice)
        deleted += 1
    db.session.commit()
    flash(f"Se eliminaron {deleted} aprendices.", "success")
    return redirect(url_for("apprentices"))


@app.route("/aprendices/exportar")
@login_required
def export_apprentices():
    apprentice_query = build_apprentice_query()
    apprentice_rows = apprentice_query.all()
    numbers = sorted({item.group_number for item in apprentice_rows if item.group_number})
    group_query = TrainingGroup.query.filter(TrainingGroup.group_number.in_(numbers or [""]))
    if current_user.role == "docente":
        managed_numbers = get_user_managed_group_numbers()
        if managed_numbers:
            group_query = group_query.filter(TrainingGroup.group_number.in_(managed_numbers))
        else:
            group_query = group_query.filter_by(created_by=current_user.id)
    group_rows = group_query.order_by(TrainingGroup.group_number.asc()).all()
    output = export_reference_workbook(apprentice_rows, group_rows)
    return send_file(
        output,
        as_attachment=True,
        download_name="gia_gestion_integral.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/aprendices/importar", methods=["POST"])
@login_required
@role_required("docente", "super_admin")
def import_apprentices():
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Debes seleccionar un archivo XLSX válido.", "error")
        return redirect(url_for("apprentices"))
    try:
        apprentice_count, group_count, has_apprentices, has_groups = import_reference_workbook(
            file,
            current_user.id,
            mode="apprentices",
        )
        if not has_apprentices and not has_groups:
            flash("No se detectó una hoja válida de aprendices en el archivo.", "error")
            return redirect(url_for("apprentices"))

        flash(
            f"Importación completada. Aprendices: {apprentice_count}.",
            "success",
        )
    except SQLAlchemyError:
        db.session.rollback()
        flash(
            "La importación no pudo completarse con la base actual. Para producción y cargas grandes usa MySQL; la app ya está preparada para ello mediante MYSQL_URL o DATABASE_URL.",
            "error",
        )
    return redirect(url_for("apprentices"))


@app.route("/fichas")
@login_required
def groups():
    if current_user.role == "aprendiz":
        return redirect(url_for("dashboard"))
    query = build_group_query()
    records = query.order_by(TrainingGroup.group_number.asc()).all()
    return render_template("groups.html", groups=records)


@app.route("/fichas/nueva", methods=["GET", "POST"])
@login_required
@role_required("docente", "super_admin")
def group_create():
    if request.method == "POST":
        data = parse_form(GROUP_FIELDS)
        group = TrainingGroup(created_by=current_user.id, **data)
        db.session.add(group)
        db.session.commit()
        flash("Ficha creada correctamente.", "success")
        return redirect(url_for("groups"))

    return render_template(
        "record_form.html",
        page_title="Nueva ficha",
        submit_label="Guardar ficha",
        fields=GROUP_FIELDS,
        record=None,
        cancel_url=url_for("groups"),
    )


@app.route("/fichas/<int:group_id>")
@login_required
def group_detail(group_id):
    group = get_group_or_404(group_id)
    return render_template(
        "detail.html",
        page_title="Detalle de la ficha",
        record=group,
        fields=GROUP_FIELDS,
        edit_url=url_for("group_edit", group_id=group.id)
        if current_user.role in {"docente", "super_admin"}
        else None,
        back_url=url_for("groups"),
    )


@app.route("/fichas/<int:group_id>/editar", methods=["GET", "POST"])
@login_required
@role_required("docente", "super_admin")
def group_edit(group_id):
    group = get_group_or_404(group_id)
    if request.method == "POST":
        data = parse_form(GROUP_FIELDS)
        for key, value in data.items():
            setattr(group, key, value)
        db.session.commit()
        flash("Ficha actualizada.", "success")
        return redirect(url_for("group_detail", group_id=group.id))

    return render_template(
        "record_form.html",
        page_title="Editar ficha",
        submit_label="Guardar cambios",
        fields=GROUP_FIELDS,
        record=group,
        cancel_url=url_for("group_detail", group_id=group.id),
    )


@app.route("/fichas/<int:group_id>/eliminar", methods=["POST"])
@login_required
@role_required("docente", "super_admin")
def group_delete(group_id):
    group = get_group_or_404(group_id)
    db.session.delete(group)
    db.session.commit()
    flash("Ficha eliminada.", "success")
    return redirect(url_for("groups"))


@app.route("/fichas/eliminar-multiples", methods=["POST"])
@login_required
@role_required("docente", "super_admin")
def group_bulk_delete():
    group_ids = request.form.getlist("selected_ids")
    if not group_ids:
        flash("Selecciona al menos una ficha para eliminar.", "error")
        return redirect(url_for("groups"))

    deleted = 0
    for group_id in group_ids:
        group = get_group_or_404(int(group_id))
        db.session.delete(group)
        deleted += 1
    db.session.commit()
    flash(f"Se eliminaron {deleted} fichas.", "success")
    return redirect(url_for("groups"))


@app.route("/fichas/exportar")
@login_required
def export_groups():
    group_query = build_group_query()
    group_rows = group_query.order_by(TrainingGroup.group_number.asc()).all()
    numbers = sorted({item.group_number for item in group_rows if item.group_number})
    apprentice_query = Apprentice.query.filter(Apprentice.group_number.in_(numbers or [""]))
    if current_user.role == "docente":
        managed_numbers = get_user_managed_group_numbers()
        if managed_numbers:
            apprentice_query = apprentice_query.filter(Apprentice.group_number.in_(managed_numbers))
        else:
            apprentice_query = apprentice_query.filter_by(created_by=current_user.id)
    elif current_user.role == "aprendiz":
        apprentice_query = apprentice_query.filter_by(student_user_id=current_user.id)
    apprentice_rows = apprentice_query.order_by(Apprentice.created_at.desc()).all()
    output = export_reference_workbook(apprentice_rows, group_rows)
    return send_file(
        output,
        as_attachment=True,
        download_name="gia_gestion_integral.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/fichas/importar", methods=["POST"])
@login_required
@role_required("docente", "super_admin")
def import_groups():
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Debes seleccionar un archivo XLSX válido.", "error")
        return redirect(url_for("groups"))
    try:
        apprentice_count, group_count, has_apprentices, has_groups = import_reference_workbook(
            file,
            current_user.id,
            mode="groups",
        )
        if not has_apprentices and not has_groups:
            flash("No se detectó una hoja válida de fichas en el archivo.", "error")
            return redirect(url_for("groups"))

        flash(
            f"Importación completada. Fichas: {group_count}.",
            "success",
        )
    except SQLAlchemyError:
        db.session.rollback()
        flash(
            "La importación no pudo completarse con la base actual. Para producción y cargas grandes usa MySQL; la app ya está preparada para ello mediante MYSQL_URL o DATABASE_URL.",
            "error",
        )
    return redirect(url_for("groups"))


@app.route("/usuarios", methods=["GET", "POST"])
@login_required
@role_required("visualizador", "super_admin")
def users():
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "docente").strip()
        email = request.form.get("email", "").strip()

        if not full_name or not username or not password:
            flash("Completa nombre, usuario y contraseña.", "error")
            return redirect(url_for("users"))

        if User.query.filter_by(username=username).first():
            flash("Ese usuario ya existe.", "error")
            return redirect(url_for("users"))

        user = User(
            username=username,
            role=role,
            full_name=full_name,
            email=email,
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash("Usuario creado correctamente.", "success")
        return redirect(url_for("users"))

    search = request.args.get("search", "").strip()
    role = request.args.get("role", "").strip()

    query = User.query
    if search:
        like = f"%{search}%"
        query = query.filter((User.username.ilike(like)) | (User.full_name.ilike(like)))
    if role:
        query = query.filter_by(role=role)

    records = query.order_by(User.created_at.desc()).all()
    return render_template("users.html", users=records)


@app.route("/usuarios/<int:user_id>/eliminar", methods=["POST"])
@login_required
@role_required("visualizador", "super_admin")
def user_delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash("No puedes eliminar tu propio usuario.", "error")
        return redirect(url_for("users"))
    db.session.delete(user)
    db.session.commit()
    flash("Usuario eliminado.", "success")
    return redirect(url_for("users"))


@app.route("/usuarios/eliminar-multiples", methods=["POST"])
@login_required
@role_required("visualizador", "super_admin")
def user_bulk_delete():
    user_ids = request.form.getlist("selected_ids")
    if not user_ids:
        flash("Selecciona al menos un usuario para eliminar.", "error")
        return redirect(url_for("users"))

    users_to_delete = User.query.filter(User.id.in_([int(item) for item in user_ids])).all()
    deleted = 0
    for user in users_to_delete:
        if user.id == current_user.id:
            continue
        db.session.delete(user)
        deleted += 1
    db.session.commit()
    if deleted:
        flash(f"Se eliminaron {deleted} usuarios.", "success")
    else:
        flash("No se eliminó ningún usuario.", "error")
    return redirect(url_for("users"))


@app.route("/bitacoras", methods=["GET", "POST"])
@login_required
def bitacoras():
    apprentices = visible_apprentices_query().order_by(Apprentice.first_names.asc(), Apprentice.last_names.asc()).all()

    if request.method == "POST":
        apprentice_id = int(request.form.get("apprentice_id"))
        apprentice = get_apprentice_or_404(apprentice_id)
        title = request.form.get("title", "").strip() or f"Bitacora {len(apprentice.bitacoras) + 1}"
        notes = request.form.get("notes", "").strip()
        uploaded_file = request.files.get("file")
        file_name = None
        file_path = None

        if uploaded_file and uploaded_file.filename:
            file_name = secure_filename(uploaded_file.filename)
            final_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file_name}"
            file_path = os.path.join(UPLOAD_DIR, final_name)
            uploaded_file.save(file_path)

        bitacora = Bitacora(
            apprentice_id=apprentice.id,
            uploaded_by_id=current_user.id,
            title=title,
            notes=notes,
            file_name=file_name,
            file_path=file_path,
        )
        db.session.add(bitacora)
        db.session.commit()
        flash("Bitácora registrada correctamente.", "success")
        return redirect(url_for("bitacoras"))

    selected_apprentice_id = request.args.get("apprentice_id", type=int)
    apprentice_search = request.args.get("apprentice_search", "").strip()
    instructor_search = request.args.get("instructor_search", "").strip()
    group_search = request.args.get("group_search", "").strip()
    selected_apprentice = None
    apprentice_ids = [item.id for item in apprentices]
    bitacora_query = Bitacora.query.join(Apprentice).filter(Apprentice.id.in_(apprentice_ids or [-1]))

    if selected_apprentice_id:
        selected_apprentice = get_apprentice_or_404(selected_apprentice_id)
        bitacora_query = bitacora_query.filter(Bitacora.apprentice_id == selected_apprentice.id)
    if apprentice_search:
        like = f"%{apprentice_search}%"
        bitacora_query = bitacora_query.filter(
            (Apprentice.first_names.ilike(like))
            | (Apprentice.last_names.ilike(like))
            | (Apprentice.document_number.ilike(like))
        )
    if instructor_search:
        like = f"%{instructor_search}%"
        bitacora_query = bitacora_query.filter(
            (Apprentice.lead_instructor.ilike(like))
            | (Apprentice.followup_instructor.ilike(like))
        )
    if group_search:
        bitacora_query = bitacora_query.filter(Apprentice.group_number.ilike(f"%{group_search}%"))

    bitacora_records = bitacora_query.order_by(Bitacora.created_at.desc()).all()

    return render_template(
        "bitacoras.html",
        apprentices=apprentices,
        bitacoras=bitacora_records,
        selected_apprentice=selected_apprentice,
    )


@app.route("/bitacoras/<int:bitacora_id>/eliminar", methods=["POST"])
@login_required
def bitacora_delete(bitacora_id):
    bitacora = Bitacora.query.get_or_404(bitacora_id)
    apprentice = get_apprentice_or_404(bitacora.apprentice_id)
    if current_user.role == "aprendiz" and apprentice.student_user_id != current_user.id:
        abort(403)
    if bitacora.file_path and os.path.exists(bitacora.file_path):
        os.remove(bitacora.file_path)
    db.session.delete(bitacora)
    db.session.commit()
    flash("Bitácora eliminada.", "success")
    return redirect(url_for("bitacoras"))


@app.route("/perfil", methods=["GET", "POST"])
@login_required
def profile():
    apprentice = None
    managed_groups = []
    available_groups = TrainingGroup.query.order_by(TrainingGroup.group_number.asc()).all() if current_user.role == "docente" else []
    if current_user.role == "aprendiz":
        apprentice = Apprentice.query.filter_by(student_user_id=current_user.id).first()
    elif current_user.role == "docente":
        managed_groups = get_user_managed_group_numbers(current_user)

    if request.method == "POST":
        current_user.full_name = request.form.get("full_name", current_user.full_name).strip()
        current_user.email = request.form.get("email", current_user.email or "").strip()
        if request.form.get("password"):
            current_user.set_password(request.form["password"].strip())
        if current_user.role == "docente":
            current_user.managed_group_numbers = "\n".join(
                normalize_group_numbers(request.form.get("managed_group_numbers", ""))
            )

        if apprentice:
            apprentice.phone = request.form.get("phone", apprentice.phone or "").strip()
            apprentice.email = current_user.email
            apprentice.municipality_origin = request.form.get(
                "municipality_origin", apprentice.municipality_origin or ""
            ).strip()
        db.session.commit()
        flash("Perfil actualizado.", "success")
        return redirect(url_for("profile"))

    return render_template(
        "profile.html",
        apprentice=apprentice,
        managed_groups=managed_groups,
        available_groups=available_groups,
    )


@app.errorhandler(403)
def forbidden(_error):
    return render_template("error.html", title="Acceso denegado", message="No tienes permisos para acceder a este módulo."), 403


@app.errorhandler(404)
def not_found(_error):
    return render_template("error.html", title="No encontrado", message="El recurso solicitado no existe o fue removido."), 404


with app.app_context():
    db.create_all()
    ensure_schema_updates()
    seed_data()


if __name__ == "__main__":
    app.run(debug=True)
