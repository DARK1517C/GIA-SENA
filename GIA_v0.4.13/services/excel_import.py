"""
services/excel_import.py

Servicio de importación de datos institucionales hacia GIA.

===============================================================================
OBJETIVO
===============================================================================

Centralizar la importación de información proveniente de las plantillas
institucionales utilizadas durante el seguimiento de la Etapa Productiva.

Este servicio permite crear y actualizar grupos y aprendices, reutilizando
el mismo flujo de negocio independientemente del formato del archivo Excel.

===============================================================================
FLUJO GENERAL
===============================================================================

1. Detectar automáticamente la plantilla utilizada.
2. Leer únicamente las hojas requeridas.
3. Normalizar la información.
4. Convertir los datos a un modelo interno común.
5. Crear o actualizar la información en GIA.
6. Generar el resultado de la importación.

===============================================================================
PLANTILLAS SOPORTADAS
===============================================================================

1. Gestión Individual Aprendices

    Hojas:

        • Aprendices
        • Record Fichas

2. Reporte Entrega Ficha a Etapa Productiva

    Hojas:

        • Control seguimiento

===============================================================================
REGLAS GENERALES
===============================================================================

• Ambas plantillas pueden crear aprendices desde cero.

• Ambas plantillas pueden actualizar información existente.

• Si un dato no viene en el Excel, nunca debe sobrescribirse la información
  existente en la plataforma.

• El aprendiz se identifica mediante el tipo y número de documento.

• La normalización de valores debe apoyarse en los catálogos del proyecto.
"""

import re
import unicodedata

from dataclasses import dataclass, field

from flask import current_app
from openpyxl import load_workbook

from .utils import (
    build_alias_lookup,
    calculate_followup_ranges,
    calculate_group_validity,
    clean_cell,
    normalize_header,
    validate_group_validity,
)
from .date_rules import audit_date_consistency, build_derived_apprentice_dates, build_derived_group_dates

# =============================================================================
# CONSTANTES
# =============================================================================

#
# Plantillas soportadas
#

TEMPLATE_GESTION_INDIVIDUAL = "gestion_individual"

TEMPLATE_REPORTE_ENTREGA = "reporte_entrega"

#
# Hojas oficiales
#

OFFICIAL_APPRENTICE_SHEET = "Aprendices"

OFFICIAL_GROUP_SHEET = "Record Fichas"

OFFICIAL_DELIVERY_SHEET = "Control seguimiento"

# =============================================================================
# DETECCIÓN DE PLANTILLAS
# =============================================================================


def detect_excel_template(workbook) -> str:
    """
    Detecta automáticamente la plantilla del libro Excel.

    Returns
    -------
    str

        TEMPLATE_GESTION_INDIVIDUAL
        TEMPLATE_REPORTE_ENTREGA

    Raises
    ------
    ValueError

        Cuando el libro no corresponde a ninguna plantilla soportada.
    """

    sheet_names = {
        normalize_header(sheet.title)
        for sheet in workbook.worksheets
    }

    has_apprentices = (
        normalize_header(OFFICIAL_APPRENTICE_SHEET)
        in sheet_names
    )

    has_groups = (
        normalize_header(OFFICIAL_GROUP_SHEET)
        in sheet_names
    )

    has_delivery = (
        normalize_header(OFFICIAL_DELIVERY_SHEET)
        in sheet_names
    )

    if has_apprentices and has_groups:
        return TEMPLATE_GESTION_INDIVIDUAL

    if has_delivery:
        return TEMPLATE_REPORTE_ENTREGA

    raise ValueError(
        "El archivo Excel no corresponde a ninguna plantilla soportada por GIA."
    )


# =============================================================================
# UTILIDADES DE NORMALIZACIÓN
# =============================================================================


def normalize_text(value):
    """
    Normaliza texto para facilitar comparaciones.

    - Convierte a minúsculas.
    - Elimina tildes.
    - Elimina espacios duplicados.
    - Unifica separadores.
    """

    if not value:
        return ""

    value = str(value).strip().lower()

    value = "".join(
        ch
        for ch in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(ch)
    )

    value = re.sub(r"[\s\-_]+", " ", value)

    return value

# =============================================================================
# NORMALIZACIONES TEMPORALES
# =============================================================================


def canonical_ep_modality(raw):
    """
    Convierte el texto proveniente de Excel al valor canónico
    aceptado por el catálogo EpModality.
    """

    key = normalize_text(raw)

    if not key:
        return None

    # -------------------------------------------------------------
    # Contrato de aprendizaje
    # -------------------------------------------------------------

    if (
        "contrato" in key
        and "aprendiz" in key
    ):
        return "CONTRATO_APRENDIZAJE"

    # -------------------------------------------------------------
    # Contrato de vínculo formativo
    # -------------------------------------------------------------

    if (
        "contrato" in key
        and "vincul" in key
        and "form" in key
    ):
        return "CONTRATO_VINCULO_FORMATIVO"

    if (
        "vincul" in key
        and "form" in key
    ):
        return "CONTRATO_VINCULO_FORMATIVO"

    # -------------------------------------------------------------
    # Vínculo laboral
    # -------------------------------------------------------------

    if (
        "vincul" in key
        and "labor" in key
    ):
        return "VINCULO_LABORAL"

    # -------------------------------------------------------------
    # Proyecto productivo
    # -------------------------------------------------------------

    if "proyect" in key:
        return "PROYECTO_PRODUCTIVO"

    # -------------------------------------------------------------
    # Prácticas en economía popular
    # -------------------------------------------------------------

    if (
        "practic" in key
        or "economia popular" in key
    ):
        return "PRACTICAS_ECONOMIA_POPULAR"

    # -------------------------------------------------------------
    # Monitoría
    # -------------------------------------------------------------

    if "monitor" in key:
        return "MONITORIA"

    return None


def canonical_yes_no(raw):
    """
    Convierte valores de Excel a los valores canónicos del catálogo YesNo.
    """

    key = normalize_text(raw)

    if not key:
        return None

    if key in {
        "si",
        "s",
        "yes",
        "true",
        "1",
        "habilitado",
        "activo",
        "continua",
        "continua en la empresa",
    }:
        return "SI"

    if key in {
        "no",
        "n",
        "false",
        "0",
        "no habilitado",
        "inactivo",
        "no continua",
        "no continua en la empresa",
    }:
        return "NO"

    return None

# =============================================================================
# RESULTADO DE IMPORTACIÓN
# =============================================================================


@dataclass
class ImportResult:
    """
    Resultado acumulado del proceso de importación.

    Esta estructura concentra toda la información generada durante la
    importación para facilitar su presentación en la interfaz y el registro
    de incidencias.
    """

    apprentice_count: int = 0
    group_count: int = 0

    has_apprentice_sheet: bool = False
    has_group_sheet: bool = False

    created_apprentices: int = 0
    updated_apprentices: int = 0

    created_groups: int = 0
    updated_groups: int = 0

    skipped_apprentices: int = 0

    errors: list[str] = field(default_factory=list)

    warnings: list[str] = field(default_factory=list)

    def __iter__(self):
        """
        Mantiene compatibilidad con el comportamiento existente.
        """

        yield self.apprentice_count
        yield self.group_count

        yield self.has_apprentice_sheet
        yield self.has_group_sheet

# =============================================================================
# ALIAS DE IMPORTACIÓN
# =============================================================================

"""
Los alias definidos en esta sección permiten que el importador reconozca
diferentes nombres de columnas provenientes de las distintas plantillas
institucionales.

Las claves representan el nombre interno utilizado por GIA.

Los valores corresponden a los encabezados aceptados durante la importación.

NOTAS
-----
- Los encabezados son normalizados antes de compararse.
- Es posible añadir nuevos alias sin modificar el resto del importador.
- Los alias de encabezados pertenecen a este importador y no forman parte del dominio de usuarios/evidencias.
"""

# -----------------------------------------------------------------------------
# Gestión Individual Aprendices
# -----------------------------------------------------------------------------

APPRENTICE_IMPORT_ALIASES = {

    # -------------------------
    # Grupo
    # -------------------------

    "group_number": [
        "N° DE FICHA",
        "Nº DE FICHA",
        "N DE FICHA",
        "FICHA",
    ],

    "lead_instructor": [
        "NOMBRE DE INSTRUCTOR(A) LIDER DE LA FICHA",
        "NOMBRE DEL INSTRUCTOR LIDER DE LA FICHA",
    ],

    "followup_instructor": [
        "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)",
        "NOMBRE DE INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO EP",
    ],

    "followup_instructor_email": [
        "CORREO DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)",
        "CORREO INSTRUCTOR SEGUIMIENTO",
    ],

    "program_name": [
        "NOMBRE DEL PROGRAMA DE FORMACION",
        "NOMBRE DEL PROGRAMA DE FORMACON",
    ],

    "program_level": [
        "NIVEL DEL PROGRAMA",
        "NIVEL DE PROGRAMA",
    ],

    # -------------------------
    # Identificación
    # -------------------------

    "document_type": [
        "TIPO DE DOCUMENTO (CC, TI, CE)",
        "TIPO DE DOCUMENTO",
    ],

    "document_number": [
        "N° DOCUMENTO DEL APRENDIZ",
        "N° DE DOCUMENTO DEL APRENDIZ",
        "N° DOCUMENTO DE IDENTIFICACION",
        "Nº DOCUMENTO DE IDENTIFICACION",
    ],

    "first_names": [
        "NOMBRES DEL APRENDIZ",
        "NOMBRES",
    ],

    "last_names": [
        "APELLIDOS DEL APRENDIZ",
        "APELLIDOS",
    ],

    "gender": [
        "GENERO (F/M)",
        "GENERO",
    ],

    "phone": [
        "N° DE CONTACTO DEL APRENDIZ",
        "TELEFONO DEL APRENDIZ",
        "TELEFONO",
    ],

    "municipality_origin": [
        "MUNICIPIO DE ORIGEN",
        "CIUDAD DE ORIGEN DEL APRENDIZ",
    ],

    "email": [
        "CORREO ELECTRONICO DEL APRENDIZ",
        "CORREO ELECTRONICO",
    ],

    # -------------------------
    # Etapa productiva
    # -------------------------

    "ep_modality": [
        "ALTERNATIVA ETAPA PRODUCTIVA",
        "MODALIDAD ETAPA PRODUCTIVA",
        "ALTERNATIVA EP",
    ],

    "practice_start_date": [
        "FECHA INICIO DE PRACTICAS",
        "FECHA DE INICIO DE EP",
    ],

    "practice_end_date": [
        "FECHA FINAL DE PRACTICAS",
        "FECHA FINAL DE EP",
    ],

    # -------------------------
    # Empresa
    # -------------------------

    "company_name": [
        "NOMBRE DE LA EMPRESA/ORG/INST",
        "NOMBRE EMPRESA",
    ],

    "company_address": [
        "DIRECCION DE LA EMPRESA",
        "DIRECCION DE EMPRESA",
    ],

    "company_municipality": [
        "MUNICIPIO",
        "MUNICIPIO EMPRESA",
    ],

    "coformador_name": [
        "NOMBRE COFORMADOR",
        "NOMBRE DEL COFORMADOR",
    ],

    "coformador_email": [
        "CORREO ELECTRONICO DEL COFORMADOR",
        "CORREO DEL COFORMADOR",
    ],

    "coformador_phone": [
        "TELEFONO DEL COFORMADOR",
    ],

    # -------------------------
    # Seguimiento
    # -------------------------

    "individual_management": [
        "GESTION INDIVIDUAL DEL APRENDIZ",
        "GESTION INDIVIDUAL DEL APRENDIZ EN EP",
    ],

    "sofia_status": [
        "ESTADO DEL APRENDIZ EN SOFIAPLUS",
        "ESTADO SOFIA PLUS",
    ],

    "arl_responsible": [
        "RESPONSABLE DE AFILIACION ARL",
        "RESPONSABLE ARL",
    ],

    "continues_company": [
        "CONTINUA EN LA EMPRESA (SI/NO)",
        "SE QUEDA LABORANDO EN LA EMPRESA",
    ],

    "evaluation_date": [
        "FECHA EMISION DE JUICIO EVALUATIVO EN SOFIA PLUS",
        "FECHA DE EVALUACION DE ETAPA PRODUCTIVA EN SOFIA PLUS",
    ],

    "english_results": [
        "JUICIOS DE INGLES APROBADOS SI/NO",
        "JUICIOS DE INGLES",
    ],
}

# -----------------------------------------------------------------------------
# Record Fichas
# -----------------------------------------------------------------------------

GROUP_IMPORT_ALIASES = {

    # -------------------------------------------------------------------------
    # Información general
    # -------------------------------------------------------------------------

    "group_number": [
        "N° DE FICHA",
        "Nº DE FICHA",
        "N DE FICHA",
        "FICHA",
        "NUMERO DE FICHA",
        "NÚMERO DE FICHA",
    ],

    "program_name": [
        "NOMBRE DEL PROGRAMA DE FORMACION",
        "NOMBRE DEL PROGRAMA DE FORMACIÓN",
        "PROGRAMA DE FORMACION",
        "PROGRAMA DE FORMACIÓN",
        "NOMBRE DEL PROGRAMA",
    ],

    "program_level": [
        "NIVEL DEL PROGRAMA",
        "NIVEL DE PROGRAMA",
        "NIVEL DE FORMACION",
        "NIVEL DE FORMACIÓN",
    ],

    "municipality": [
        "MUNICIPIO",
        "MUNICIPIO DE LA FICHA",
        "MUNICIPIO CENTRO",
        "CIUDAD",
    ],

    "modality": [
        "MODALIDAD",
        "MODALIDAD DE FORMACION",
        "MODALIDAD DE FORMACIÓN",
        "TIPO DE MODALIDAD",
    ],

    # -------------------------------------------------------------------------
    # Instructores
    # -------------------------------------------------------------------------

    "lead_instructor": [
        "NOMBRE DE INSTRUCTOR(A) LIDER DE LA FICHA",
        "NOMBRE DEL INSTRUCTOR LIDER DE LA FICHA",
        "INSTRUCTOR LIDER",
        "INSTRUCTOR LÍDER",
        "INSTRUCTOR(A) LIDER",
        "INSTRUCTOR(A) LÍDER",
    ],

    "followup_instructor": [
        "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO ETAPA PRODUCTIVA (EP)",
        "NOMBRE DE INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "NOMBRE DE INSTRUCTOR(A) DE SEGUIMIENTO EP",
        "INSTRUCTOR DE SEGUIMIENTO",
        "INSTRUCTOR(A) DE SEGUIMIENTO",
        "INSTRUCTOR SEGUIMIENTO",
    ],

    # -------------------------------------------------------------------------
    # Estado
    # -------------------------------------------------------------------------

    "sofia_group_status": [
        "ESTADO DE LA FICHA EN SOFIA PLUS",
        "ESTADO DE LA FICHA EN SOFIAPLUS",
        "ESTADO FICHA SOFIA PLUS",
        "ESTADO FICHA SOFIAPLUS",
        "ESTADO SOFIA PLUS",
        "ESTADO SOFIAPLUS",
    ],

    "group_validity": [
        "VIGENCIA DE LA FICHA",
        "VIGENCIA FICHA",
        "ESTADO DE VIGENCIA",
        "VIGENCIA",
    ],

    # -------------------------------------------------------------------------
    # Fechas
    # -------------------------------------------------------------------------

    "group_start_date": [
        "FECHA DE INICIO DE LA FICHA",
        "FECHA INICIO DE LA FICHA",
        "FECHA INICIO FICHA",
        "FECHA DE INICIO",
    ],

    "training_end_date": [
        "FECHA DE TERMINACION DE LA FORMACION",
        "FECHA DE TERMINACIÓN DE LA FORMACIÓN",
        "FECHA TERMINACION FORMACION",
        "FECHA TERMINACIÓN FORMACIÓN",
        "FECHA FIN DE FORMACION",
        "FECHA FIN DE FORMACIÓN",
        "FECHA FIN",
    ],

    "ep_start_date": [
        "FECHA DE INICIO DE ETAPA PRODUCTIVA",
        "FECHA INICIO ETAPA PRODUCTIVA",
        "FECHA INICIO EP",
        "INICIO ETAPA PRODUCTIVA",
        "INICIO EP",
    ],

    # -------------------------------------------------------------------------
    # Estadísticas
    # -------------------------------------------------------------------------

    "apprentices_training": [
        "APRENDICES EN FORMACION",
        "APRENDICES EN FORMACIÓN",
        "APRENDICES FORMACION",
        "APRENDICES FORMACIÓN",
    ],

    "apprentices_enabled": [
        "APRENDICES ACTIVOS",
        "APRENDICES HABILITADOS",
        "APRENDICES VIGENTES",
    ],

    "apprentices_rap_pending": [
        "APRENDICES RAP PENDIENTES",
        "APRENDICES RAP PENDIENTE",
        "RAP PENDIENTES",
        "RAP PENDIENTE",
    ],

    "apprentices_practice": [
        "APRENDICES EN PRACTICA",
        "APRENDICES EN PRÁCTICA",
        "APRENDICES EN ETAPA PRODUCTIVA",
        "APRENDICES ETAPA PRODUCTIVA",
    ],

    "apprentices_without_alternative": [
        "APRENDICES SIN ALTERNATIVA",
        "APRENDICES SIN ALTERNATIVA DE ETAPA PRODUCTIVA",
        "SIN ALTERNATIVA",
    ],

    "apprentices_certified": [
        "APRENDICES CERTIFICADOS",
        "APRENDICES CERTIFICADO",
        "CERTIFICADOS",
    ],

    # -------------------------------------------------------------------------
    # Modalidades de etapa productiva
    # -------------------------------------------------------------------------

    "learning_contract": [
        "CONTRATO DE APRENDIZAJE",
        "CONTRATOS DE APRENDIZAJE",
        "APRENDICES CONTRATO DE APRENDIZAJE",
    ],

    "internship": [
        "VINCULO FORMATIVO",
        "VÍNCULO FORMATIVO",
    ],

    "productive_project": [
        "PROYECTO PRODUCTIVO",
        "PROYECTOS PRODUCTIVOS",
    ],

    "employment_link": [
        "VINCULACION LABORAL",
        "VINCULACIÓN LABORAL",
        "VINCULO LABORAL",
        "VÍNCULO LABORAL",
        "VINCULACION",
        "VINCULACIÓN",
    ],
}

DELIVERY_IMPORT_ALIASES = {
    "group_number": [
        "N° DE FICHA",
        "N DE FICHA",
        "NUMERO DE FICHA",
        "NÚMERO DE FICHA",
    ],

    "lead_instructor": [
        "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
        "NOMBRE DE INSTRUCTOR LÍDER DE LA FICHA",
        "INSTRUCTOR LÍDER",
    ],

    "followup_instructor": [
        "NOMBRE DE INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "NOMBRE DE INSTRUCTOR ETAPA PRODUCTIVA",
        "INSTRUCTOR ETAPA PRODUCTIVA",
    ],

    "program_name": [
        "NOMBRE DEL PROGRAMA DE FORMACIÓN",
        "NOMBRE DEL PROGRAMA DE FORMACION",
        "PROGRAMA DE FORMACIÓN",
        "PROGRAMA DE FORMACION",
    ],

    "program_level": [
        "NIVEL DEL PROGRAMA",
        "NIVEL DE FORMACIÓN",
        "NIVEL DE FORMACION",
    ],

    "document_type": [
        "TIPO DE DOCUMENTO (CC, TI, CE)",
        "TIPO DE DOCUMENTO",
    ],

    "document_number": [
        "N° DOCUMENTO DEL APRENDIZ",
        "N DOCUMENTO DEL APRENDIZ",
        "NUMERO DOCUMENTO DEL APRENDIZ",
        "DOCUMENTO DEL APRENDIZ",
    ],

    "first_names": [
        "NOMBRES DEL APRENDIZ",
        "NOMBRES",
    ],

    "last_names": [
        "APELLIDOS DEL APRENDIZ",
        "APELLIDOS",
    ],

    "gender": [
        "GENERO",
        "GÉNERO",
    ],

    "phone": [
        "N° DE CONTACTO DEL APRENDIZ",
        "N DE CONTACTO DEL APRENDIZ",
        "NUMERO DE CONTACTO DEL APRENDIZ",
        "CONTACTO DEL APRENDIZ",
    ],

    "email": [
        "CORREO ELECTRÓNICO DEL APRENDIZ",
        "CORREO ELECTRONICO DEL APRENDIZ",
        "CORREO DEL APRENDIZ",
    ],

    "ep_modality": [
        "ALTERNATIVA ETAPA PRODUCTIVA",
        "ALTERNATIVA DE ETAPA PRODUCTIVA",
        "MODALIDAD ETAPA PRODUCTIVA",
    ],

    "practice_start_date": [
        "FECHA INICIO DE PRÁCTICAS",
        "FECHA INICIO DE PRACTICAS",
        "INICIO DE PRÁCTICAS",
        "INICIO DE PRACTICAS",
    ],

    "practice_end_date": [
        "FECHA FINAL DE PRÁCTICAS",
        "FECHA FINAL DE PRACTICAS",
        "FECHA FIN DE PRÁCTICAS",
        "FECHA FIN DE PRACTICAS",
        "FINAL DE PRÁCTICAS",
        "FINAL DE PRACTICAS",
    ],

    "company_name": [
        "NOMBRE DE LA EMPRESA/ORG/INST",
        "NOMBRE DE LA EMPRESA",
        "EMPRESA/ORG/INST",
        "EMPRESA",
    ],

    "company_municipality": [
        "MUNICIPIO",
        "MUNICIPIO EMPRESA",
        "MUNICIPIO DE LA EMPRESA",
    ],

    "coformador_name": [
        "NOMBRE COFORMADOR",
        "NOMBRE DEL COFORMADOR",
        "COFORMADOR",
    ],

    "coformador_email": [
        "CORREO ELECTRÓNICO DEL COFORMADOR",
        "CORREO ELECTRONICO DEL COFORMADOR",
        "CORREO DEL COFORMADOR",
    ],

    "coformador_phone": [
        "TELÉFONO DEL COFORMADOR",
        "TELEFONO DEL COFORMADOR",
        "TELÉFONO COFORMADOR",
        "TELEFONO COFORMADOR",
    ],

    "followup_moments": [
        "OBSERVACIONES INSTRUCTOR(A) TÉCNICO",
        "OBSERVACIONES INSTRUCTOR(A) TECNICO",
        "OBSERVACIONES INSTRUCTOR TÉCNICO",
        "OBSERVACIONES INSTRUCTOR TECNICO",
        "OBSERVACIONES",
    ],

    "individual_management": [
        "GESTIÓN INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "GESTION INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "GESTIÓN INSTRUCTOR ETAPA PRODUCTIVA",
        "GESTION INSTRUCTOR ETAPA PRODUCTIVA",
    ],
}

# =============================================================================
# CAMPOS DEL MODELO APPRENTICE
# =============================================================================

APPRENTICE_MODEL_FIELDS = [

    # Grupo
    "group_id",
    "group_number",
    "program_name",
    "program_level",
    "group_validity",

    # Identificación
    "document_type",
    "document_number",
    "first_names",
    "last_names",
    "gender",
    "phone",
    "email",
    "municipality_origin",

    # Instructores
    "lead_instructor",
    "followup_instructor",
    "followup_instructor_email",

    # Etapa productiva
    "ep_modality",
    "sofia_status",
    "practice_start_date",
    "practice_end_date",

    # Rangos de seguimiento
    "followup_moment1_start",
    "followup_moment1_end",
    "followup_moment2_start",
    "followup_moment2_end",
    "followup_moment3_start",
    "followup_moment3_end",
    "followup_moment4_start",
    "followup_moment4_end",

    # Empresa
    "company_name",
    "company_municipality",
    "company_address",

    # Coformador
    "coformador_name",
    "coformador_email",
    "coformador_phone",

    # Seguimiento
    "arl_responsible",
    "continues_company",
    "individual_management",
    "followup_moments",
    "evaluation_date",
    "english_results",
]

# =============================================================================
# CAMPOS DEL MODELO TRAINING GROUP
# =============================================================================

GROUP_MODEL_FIELDS = [

    # Información general
    "group_number",
    "program_name",
    "program_level",
    "municipality",
    "modality",

    # Instructores
    "lead_instructor",
    "followup_instructor",

    # Estado
    "sofia_group_status",
    "group_validity",

    # Fechas
    "group_start_date",
    "training_end_date",
    "ep_start_date",

    # Estadísticas
    "apprentices_statistics",
    "apprentices_training",
    "apprentices_enabled",
    "apprentices_rap_pending",
    "apprentices_practice",
    "apprentices_without_alternative",
    "apprentices_certified",

    # Modalidades
    "productive_modalities",
    "learning_contract",
    "internship",
    "productive_project",
    "employment_link",
]

# =============================================================================
# TABLAS DE BÚSQUEDA DE ALIAS
# =============================================================================

APPRENTICE_ALIAS_LOOKUP = build_alias_lookup(
    APPRENTICE_IMPORT_ALIASES
)

DELIVERY_ALIAS_LOOKUP = build_alias_lookup(
    DELIVERY_IMPORT_ALIASES
)

GROUP_ALIAS_LOOKUP = build_alias_lookup(
    GROUP_IMPORT_ALIASES
)

DELIVERY_IMPORT_ALIASES = {
    "group_number": [
        "N° DE FICHA",
        "N DE FICHA",
        "NUMERO DE FICHA",
        "NÚMERO DE FICHA",
    ],

    "lead_instructor": [
        "NOMBRE DE INSTRUCTOR(A) LÍDER DE LA FICHA",
        "NOMBRE DE INSTRUCTOR LÍDER DE LA FICHA",
        "INSTRUCTOR(A) LÍDER DE LA FICHA",
    ],

    "followup_instructor": [
        "NOMBRE DE INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "NOMBRE DE INSTRUCTOR ETAPA PRODUCTIVA",
        "INSTRUCTOR(A) ETAPA PRODUCTIVA",
    ],

    "program_name": [
        "NOMBRE DEL PROGRAMA DE FORMACIÓN",
        "NOMBRE DEL PROGRAMA DE FORMACION",
        "PROGRAMA DE FORMACIÓN",
        "PROGRAMA DE FORMACION",
    ],

    "program_level": [
        "NIVEL DEL PROGRAMA",
    ],

    "document_type": [
        "TIPO DE DOCUMENTO (CC, TI, CE)",
        "TIPO DE DOCUMENTO",
    ],

    "document_number": [
        "N° DOCUMENTO DEL APRENDIZ",
        "N DE DOCUMENTO DEL APRENDIZ",
        "NUMERO DE DOCUMENTO DEL APRENDIZ",
        "DOCUMENTO DEL APRENDIZ",
    ],

    "first_names": [
        "NOMBRES DEL APRENDIZ",
        "NOMBRES",
    ],

    "last_names": [
        "APELLIDOS DEL APRENDIZ",
        "APELLIDOS",
    ],

    "gender": [
        "GENERO",
        "GÉNERO",
    ],

    "phone": [
        "N° DE CONTACTO DEL APRENDIZ",
        "N DE CONTACTO DEL APRENDIZ",
        "NUMERO DE CONTACTO DEL APRENDIZ",
        "CONTACTO DEL APRENDIZ",
    ],

    "email": [
        "CORREO ELECTRÓNICO DEL APRENDIZ",
        "CORREO ELECTRONICO DEL APRENDIZ",
        "EMAIL DEL APRENDIZ",
    ],

    "ep_modality": [
        "ALTERNATIVA ETAPA PRODUCTIVA",
        "ALTERNATIVA DE ETAPA PRODUCTIVA",
    ],

    "practice_start_date": [
        "FECHA INICIO DE PRÁCTICAS",
        "FECHA INICIO DE PRACTICAS",
        "FECHA INICIO",
    ],

    "practice_end_date": [
        "FECHA FINAL DE PRÁCTICAS",
        "FECHA FINAL DE PRACTICAS",
        "FECHA FIN DE PRÁCTICAS",
        "FECHA FIN DE PRACTICAS",
    ],

    "company_name": [
        "NOMBRE DE LA EMPRESA/ORG/INST",
        "NOMBRE DE LA EMPRESA",
        "EMPRESA/ORG/INST",
    ],

    "company_municipality": [
        "MUNICIPIO",
        "MUNICIPIO EMPRESA",
    ],

    "coformador_name": [
        "NOMBRE COFORMADOR",
        "NOMBRE DEL COFORMADOR",
    ],

    "coformador_email": [
        "CORREO ELECTRÓNICO DEL COFORMADOR",
        "CORREO ELECTRONICO DEL COFORMADOR",
        "EMAIL DEL COFORMADOR",
    ],

    "coformador_phone": [
        "TELÉFONO DEL COFORMADOR",
        "TELEFONO DEL COFORMADOR",
        "N° TELÉFONO DEL COFORMADOR",
    ],

    "individual_management": [
        "OBSERVACIONES INSTRUCTOR(A) TÉCNICO",
        "OBSERVACIONES INSTRUCTOR(A) TECNICO",
        "OBSERVACIONES INSTRUCTOR TÉCNICO",
    ],

    "continues_company": [
        "GESTIÓN INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "GESTION INSTRUCTOR(A) ETAPA PRODUCTIVA",
        "GESTIÓN INSTRUCTOR ETAPA PRODUCTIVA",
        "GESTION INSTRUCTOR ETAPA PRODUCTIVA",
    ],
}

DELIVERY_ALIAS_LOOKUP = build_alias_lookup(
    DELIVERY_IMPORT_ALIASES
)

# =============================================================================
# UTILIDADES PARA HOJAS EXCEL
# =============================================================================

# =============================================================================
# CAMPOS DEL MODELO APPRENTICE
# =============================================================================

APPRENTICE_MODEL_FIELDS = [

    # Grupo
    "group_id",
    "group_number",
    "program_name",
    "program_level",
    "group_validity",

    # Identificación
    "document_type",
    "document_number",
    "first_names",
    "last_names",
    "gender",
    "phone",
    "email",
    "municipality_origin",

    # Instructores
    "lead_instructor",
    "followup_instructor",
    "followup_instructor_email",

    # Etapa productiva
    "ep_modality",
    "sofia_status",
    "practice_start_date",
    "practice_end_date",

    # Rangos de seguimiento
    "followup_moment1_start",
    "followup_moment1_end",
    "followup_moment2_start",
    "followup_moment2_end",
    "followup_moment3_start",
    "followup_moment3_end",
    "followup_moment4_start",
    "followup_moment4_end",

    # Empresa
    "company_name",
    "company_municipality",
    "company_address",

    # Coformador
    "coformador_name",
    "coformador_email",
    "coformador_phone",

    # Seguimiento
    "arl_responsible",
    "continues_company",
    "individual_management",
    "followup_moments",
    "evaluation_date",
    "english_results",
]

# =============================================================================
# CAMPOS DEL MODELO TRAINING GROUP
# =============================================================================

GROUP_MODEL_FIELDS = [

    # Información general
    "group_number",
    "program_name",
    "program_level",
    "municipality",
    "modality",

    # Instructores
    "lead_instructor",
    "followup_instructor",

    # Estado
    "sofia_group_status",
    "group_validity",

    # Fechas
    "group_start_date",
    "training_end_date",
    "ep_start_date",

    # Estadísticas
    "apprentices_statistics",
    "apprentices_training",
    "apprentices_enabled",
    "apprentices_rap_pending",
    "apprentices_practice",
    "apprentices_without_alternative",
    "apprentices_certified",

    # Modalidades
    "productive_modalities",
    "learning_contract",
    "internship",
    "productive_project",
    "employment_link",
]

# =============================================================================
# TABLAS DE BÚSQUEDA DE ALIAS
# =============================================================================

APPRENTICE_ALIAS_LOOKUP = build_alias_lookup(
    APPRENTICE_IMPORT_ALIASES
)

GROUP_ALIAS_LOOKUP = build_alias_lookup(
    GROUP_IMPORT_ALIASES
)

def _sheet_by_official_name(workbook, official_name):
    """
    Busca una hoja utilizando su nombre oficial, ignorando diferencias de
    mayúsculas, tildes y espacios.

    Parameters
    ----------
    workbook
        Libro de Excel abierto.

    official_name
        Nombre oficial esperado.

    Returns
    -------
    Worksheet | None
    """

    wanted = normalize_header(official_name)

    for sheet in workbook.worksheets:

        if normalize_header(sheet.title) == wanted:
            return sheet

    return None


# =============================================================================
# UTILIDADES DE PARSEO
# =============================================================================

def _header_map(
    sheet,
    header_row,
    subheader_row=None,
    alias_lookup=None,
):
    """
    Construye el mapa:

        columna Excel -> campo interno de GIA

    utilizando la tabla de alias correspondiente.
    """

    mapping = {}

    alias_lookup = alias_lookup or {}

    for col in range(1, sheet.max_column + 1):

        candidates = [
            sheet.cell(header_row, col).value,
        ]

        if subheader_row:
            candidates.insert(
                0,
                sheet.cell(subheader_row, col).value,
            )

        for value in candidates:

            normalized = normalize_header(value)

            if normalized and normalized in alias_lookup:
                mapping[col] = alias_lookup[normalized]
                break

    return mapping


# =============================================================================
# PARSEO DE PLANTILLAS
# =============================================================================

# -----------------------------------------------------------------------------
# Gestión Individual Aprendices
# -----------------------------------------------------------------------------

def parse_apprentice_sheet(sheet):
    """
    Convierte la hoja 'Aprendices' al modelo interno utilizado por GIA.
    """

    header_map = _header_map(
        sheet,
        1,
        alias_lookup=APPRENTICE_ALIAS_LOOKUP,
    )

    records = {}

    for row_index in range(2, sheet.max_row + 1):

        record = {
            key: ""
            for key in APPRENTICE_MODEL_FIELDS
        }

        has_data = False

        #
        # Lectura de columnas
        #

        for col_index, field in header_map.items():

            value = clean_cell(
                sheet.cell(row_index, col_index).value
            )

            if value:
                has_data = True

            if field in record:
                record[field] = value

        #
        # Momentos de seguimiento
        #

        moments = []

        for col_index in range(18, 22):

            value = clean_cell(
                sheet.cell(row_index, col_index).value
            )

            if value:
                moments.append(value)

        while len(moments) < 4:
            moments.append("")

        record["followup_moments"] = " | ".join(
            moments
        ).strip(" |")

        #
        # Rangos automáticos
        #

        record = build_derived_apprentice_dates(record)

        #
        # Registro único
        #

        if has_data and record.get("document_number"):
            records[
                record["document_number"]
            ] = record

    return list(records.values())


# -----------------------------------------------------------------------------
# Record Fichas
# -----------------------------------------------------------------------------

def parse_group_sheet(
    sheet,
    result=None,
):
    """
    Convierte la hoja 'Record Fichas' al modelo interno de grupos.
    """

    header_map = _header_map(
        sheet,
        2,
        subheader_row=3,
        alias_lookup=GROUP_ALIAS_LOOKUP,
    )

    records = {}

    for row_index in range(4, sheet.max_row + 1):

        record = {
            key: ""
            for key in GROUP_MODEL_FIELDS
        }

        has_data = False

        #
        # Lectura
        #

        for col_index, field in header_map.items():

            value = clean_cell(
                sheet.cell(row_index, col_index).value
            )

            if value:
                has_data = True

            if field in record:
                record[field] = value

        #
        # Vigencia automática
        #

        if not record.get("group_validity"):

            record["group_validity"] = (
                calculate_group_validity(
                    record.get("training_end_date")
                )
            )

        #
        # Validaciones
        #

        notes = validate_group_validity(
            record.get("ep_start_date"),
            record.get("training_end_date"),
            record.get("group_validity"),
        )

        if notes and result:

            result.warnings.append(
                f"Ficha {record.get('group_number')}: "
                + " ".join(notes)
            )

        #
        # Modalidades
        #

        modality_parts = []

        for key, label in [

            ("learning_contract", "Contrato de aprendizaje"),

            ("internship", "Vínculo formativo"),

            ("employment_link", "Vinculación laboral"),

            ("productive_project", "Proyecto productivo"),

        ]:

            if record.get(key):

                modality_parts.append(
                    f"{label}: {record[key]}"
                )

        record["productive_modalities"] = " | ".join(
            modality_parts
        )

        #
        # Estadísticas
        #

        total_stats = [

            record.get("apprentices_training"),

            record.get("apprentices_enabled"),

            record.get("apprentices_rap_pending"),

            record.get("apprentices_practice"),

            record.get("apprentices_certified"),

        ]

        record["apprentices_statistics"] = " / ".join(

            value

            for value in total_stats

            if value

        )

        #
        # Registro único
        #

        if has_data and record.get("group_number"):

            records[
                record["group_number"]
            ] = record

    return list(records.values())


# -----------------------------------------------------------------------------
# Reporte Entrega Ficha a Etapa Productiva
# -----------------------------------------------------------------------------

def parse_delivery_sheet(sheet):
    """
    Convierte la hoja ``Control seguimiento`` en dos colecciones:

    - aprendices: campos provenientes de APPRENTICE_IMPORT_ALIASES;
    - grupos: campos provenientes de GROUP_IMPORT_ALIASES.

    La fila 3 de la plantilla contiene las equivalencias internas definidas
    por el usuario. Una celda puede contener más de una equivalencia separada
    por salto de línea; por ejemplo ``practice_start_date`` y ``ep_start_date``.

    La fila 5 contiene los encabezados visibles y la fila 6 en adelante los
    datos. La columna cuya equivalencia sea ``SE OMITE`` no se importa.
    """

    apprentice_fields = set(APPRENTICE_MODEL_FIELDS)
    group_fields = set(GROUP_MODEL_FIELDS)

    # columna Excel -> campos internos. Se lee la fila 3, no los encabezados
    # institucionales de la fila 5, porque la fila 3 es la definición explícita
    # de equivalencias de esta plantilla.
    field_map = {}

    for col_index in range(1, sheet.max_column + 1):
        raw_mapping = sheet.cell(3, col_index).value

        if not raw_mapping:
            continue

        candidates = [
            str(value).strip()
            for value in str(raw_mapping).splitlines()
            if str(value).strip()
        ]

        fields = []

        for candidate in candidates:
            normalized = normalize_header(candidate)

            if normalized in ("se omite", "omite"):
                continue

            # La fila 3 contiene nombres internos, no encabezados visibles.
            # normalize_header permite tolerar espacios/tildes, pero el campo
            # final debe pertenecer a uno de los modelos conocidos.
            for field in (candidate, normalized):
                if field in apprentice_fields or field in group_fields:
                    if field not in fields:
                        fields.append(field)
                    break

        if fields:
            field_map[col_index] = fields

    apprentices = {}
    groups = {}

    for row_index in range(6, sheet.max_row + 1):
        apprentice_record = {
            key: ""
            for key in APPRENTICE_MODEL_FIELDS
        }
        group_record = {
            key: ""
            for key in GROUP_MODEL_FIELDS
        }

        apprentice_has_data = False
        group_has_data = False

        for col_index, fields in field_map.items():
            value = clean_cell(
                sheet.cell(row_index, col_index).value
            )

            if not value:
                continue

            for field in fields:
                if field in apprentice_record:
                    apprentice_record[field] = value
                    apprentice_has_data = True

                if field in group_record:
                    group_record[field] = value
                    group_has_data = True

        if not apprentice_has_data and not group_has_data:
            continue

        document_number = (
            apprentice_record.get("document_number") or ""
        ).strip()
        group_number = (
            group_record.get("group_number") or ""
        ).strip()

        # -------------------------------------------------------------
        # GRUPO
        # -------------------------------------------------------------
        if group_number:
            group_record = build_derived_group_dates(group_record)

            # Si la plantilla contiene la misma ficha varias veces, acumulamos
            # la última fila que aporte información a cada campo.
            previous = groups.get(group_number)
            if previous:
                for key, value in group_record.items():
                    if value not in (None, ""):
                        previous[key] = value
            else:
                groups[group_number] = group_record

        # -------------------------------------------------------------
        # APRENDIZ
        # -------------------------------------------------------------
        if document_number:
            # EpModality debe llegar al modelo con el valor del catálogo,
            # no con la etiqueta de presentación.
            raw_ep = apprentice_record.get("ep_modality")
            canon_ep = canonical_ep_modality(raw_ep)
            apprentice_record["ep_modality"] = canon_ep

            if raw_ep and not canon_ep:
                current_app.logger.warning(
                    "Import delivery: modalidad EP no reconocida en fila %s: %r. Se omitirá.",
                    row_index,
                    raw_ep,
                )

            # YesNo usa SI/NO en el catálogo. La plantilla puede traer
            # valores como Habilitado/No habilitado.
            for field in ("continues_company", "individual_management"):
                raw_value = apprentice_record.get(field)
                if raw_value:
                    normalized_value = canonical_yes_no(raw_value)
                    apprentice_record[field] = normalized_value

                    if normalized_value is None:
                        current_app.logger.warning(
                            "Import delivery: valor YesNo no reconocido para %s en fila %s: %r. Se omitirá.",
                            field,
                            row_index,
                            raw_value,
                        )

            apprentice_record = build_derived_apprentice_dates(apprentice_record)

            # No se importa la columna OBSERVACIONES INSTRUCTOR(A) TÉCNICO.
            # En la fila 3 aparece como SE OMITE, por lo que nunca entra en
            # field_map.
            apprentice_record["group_id"] = None

            apprentices[document_number] = apprentice_record

    return list(apprentices.values()), list(groups.values())


# =============================================================================
# CREACIÓN Y ACTUALIZACIÓN DE USUARIOS
# =============================================================================

def upsert_student_user(
    apprentice,
    known_users=None,
):
    """
    Crea o actualiza el usuario asociado a un aprendiz.

    El usuario se identifica mediante el número de documento y siempre
    pertenece al rol canónico APPRENTICE.

    Parameters
    ----------
    apprentice
        Diccionario o instancia Apprentice.

    known_users
        Caché opcional de usuarios ya cargados para evitar consultas
        repetidas a la base de datos.

    Returns
    -------
    int | None

        ID del usuario creado o actualizado.
    """

    from extensions import db
    from catalogs.user import UserDocumentType, UserRole
    from models import User

    if known_users is None:

        known_users = {
            user.document_number: user
            for user in User.query.filter_by(role=UserRole.APPRENTICE.value).all()
        }

    if isinstance(apprentice, dict):

        document_number = apprentice.get("document_number")

        first_names = apprentice.get("first_names", "")

        last_names = apprentice.get("last_names", "")

        email = apprentice.get("email")

    else:

        document_number = apprentice.document_number

        first_names = apprentice.first_names

        last_names = apprentice.last_names

        email = apprentice.email

    if not document_number:
        return None

    user = known_users.get(document_number)

    if user is None:
        user = User.query.filter_by(
            document_number=document_number
        ).first()

    if user is None:

        user = User(
            document_type=UserDocumentType.NATIONAL_ID.value,
            document_number=document_number,
            first_names=first_names,
            last_names=last_names,
            role=UserRole.APPRENTICE.value,
            email=email,
        )

        #
        # Contraseña inicial
        #

        user.set_password(document_number)

        db.session.add(user)

        db.session.flush()

        known_users[document_number] = user

    else:

        full_name = f"{first_names} {last_names}".strip()

        if first_names or last_names:
            user.first_names = first_names
            user.last_names = last_names

        if email:
            user.email = email

    if isinstance(apprentice, dict):

        return user.id

    apprentice.student_user_id = user.id

    return user.id


# =============================================================================
# CARGA DE PLANTILLAS
# =============================================================================

def _load_template(file_storage):
    """
    Abre el archivo Excel, detecta automáticamente la plantilla y
    devuelve los registros convertidos al modelo interno utilizado
    por GIA.

    Returns
    -------
    tuple

        (
            template,
            apprentices,
            groups,
        )
    """

    workbook = load_workbook(
        file_storage,
        data_only=False,
    )

    template = detect_excel_template(
        workbook,
    )

    # -------------------------------------------------------------------------
    # Gestión Individual Aprendices
    # -------------------------------------------------------------------------

    if template == TEMPLATE_GESTION_INDIVIDUAL:

        apprentice_sheet = _sheet_by_official_name(
            workbook,
            OFFICIAL_APPRENTICE_SHEET,
        )

        group_sheet = _sheet_by_official_name(
            workbook,
            OFFICIAL_GROUP_SHEET,
        )

        apprentices = parse_apprentice_sheet(
            apprentice_sheet,
        )

        groups = parse_group_sheet(
            group_sheet,
        )

    # -------------------------------------------------------------------------
    # Reporte Entrega Ficha a Etapa Productiva
    # -------------------------------------------------------------------------

    elif template == TEMPLATE_REPORTE_ENTREGA:

        delivery_sheet = _sheet_by_official_name(
            workbook,
            OFFICIAL_DELIVERY_SHEET,
        )

        if delivery_sheet is None:
            raise RuntimeError(
                f"La plantilla {TEMPLATE_REPORTE_ENTREGA!r} "
                f"no contiene la hoja "
                f"{OFFICIAL_DELIVERY_SHEET!r}."
        )

        apprentices, groups = parse_delivery_sheet(
            delivery_sheet,
        )

    else:

        raise RuntimeError(
            "Plantilla de Excel no soportada."
        )

    return (
        template,
        apprentices,
        groups,
    )

# =============================================================================
# IMPORTACIÓN PRINCIPAL
# =============================================================================

def import_reference_workbook(file_storage, owner_id, mode="both", group_scope=None):
    from extensions import db
    from catalogs.user import UserRole

    from models import (
        Apprentice,
        TrainingGroup,
        User,
    )

    from services.evidence_service import (
        ensure_submissions_for_apprentice,
        ensure_template_activities_for_group,
        sync_group_followup_dates,
    )

    try:
        # -------------------------------------------------------------------------
        # Leer plantilla
        # -------------------------------------------------------------------------
    
        template, apprentices, groups = _load_template(
            file_storage,
        )
    
        result = ImportResult()
    
        # -------------------------------------------------------------------------
        # Caché de objetos existentes
        # -------------------------------------------------------------------------
    
        existing_groups = {
            group.group_number: group
            for group in TrainingGroup.query.all()
        }
    
        existing_apprentices = {
            apprentice.document_number: apprentice
            for apprentice in Apprentice.query.all()
        }
    
        known_users = {
            user.document_number: user
            for user in User.query.filter_by(
                role=UserRole.APPRENTICE.value
            ).all()
        }
    
        # -------------------------------------------------------------------------
        # Procesar grupos
        # -------------------------------------------------------------------------
    
        if groups:
    
            for index, data in enumerate(groups, start=1):
    
                # Solo actualizar campos realmente diligenciados.
                # Un Excel parcial no debe borrar información existente en GIA.
                clean_data = {
                    key: value
                    for key, value in data.items()
                    if (
                        key in GROUP_MODEL_FIELDS
                        and value not in (None, "")
                    )
                }
    
                group_obj = existing_groups.get(
                    data["group_number"]
                )

                # Seguridad 3.D: las importaciones masivas respetan el
                # alcance del usuario y no pueden mutar fichas ajenas.
                if group_scope is not None and not group_scope(group_obj, data):
                    result.errors.append(
                        f"Fila de ficha {index} omitida: fuera del alcance del usuario."
                    )
                    continue
    
                if group_obj is None:
    
                    group_obj = TrainingGroup(
                        created_by=owner_id,
                        **clean_data,
                    )
    
                    db.session.add(group_obj)
    
                    db.session.flush()
    
                    existing_groups[
                        data["group_number"]
                    ] = group_obj
    
                    result.created_groups += 1
    
                else:
    
                    for key, value in clean_data.items():
    
                        if hasattr(group_obj, key):
                            setattr(group_obj, key, value)
    
                    group_obj.created_by = owner_id
    
                    result.updated_groups += 1
    
                ensure_template_activities_for_group(
                    group_obj
                )
    
                sync_group_followup_dates(
                    group_obj
                )
    
                result.group_count += 1
    
            # -------------------------------------------------------------
            # Normalizar catálogos provenientes de plantillas antiguas
            # -------------------------------------------------------------
    
            yes_no_fields = (
                "continues_company",
                "individual_management",
            )
    
            for field in yes_no_fields:
    
                raw_value = clean_data.get(field)
    
                if not isinstance(raw_value, str):
                    continue
    
                normalized = normalize_text(raw_value)
    
                if normalized in ("si", "s", "habilitado", "habilitada"):
                    clean_data[field] = "SI"
    
                elif normalized in ("no", "n", "deshabilitado", "deshabilitada"):
                    clean_data[field] = "NO"
    
        # -------------------------------------------------------------------------
        # Procesar aprendices
        # -------------------------------------------------------------------------
    
        if apprentices:
    
            for index, data in enumerate(
                apprentices,
                start=1,
            ):
    
                group_number = data.get(
                    "group_number"
                )
    
                group_obj = existing_groups.get(
                    group_number
                )

                # Seguridad 3.D: una hoja de aprendices tampoco puede
                # crear/modificar una ficha fuera del alcance.
                if group_scope is not None and not group_scope(group_obj, data):
                    result.errors.append(
                        f"Fila de aprendiz {index} omitida: ficha fuera del alcance del usuario."
                    )
                    result.skipped_apprentices += 1
                    continue
    
                #
                # Crear grupo provisional
                #
    
                if not group_obj:
    
                    group_obj = TrainingGroup(
    
                        group_number=group_number,
    
                        program_name=data.get(
                            "program_name",
                            "SIN DEFINIR",
                        ),
    
                        program_level=data.get(
                            "program_level",
                            "",
                        ),
    
                        lead_instructor=data.get(
                            "lead_instructor",
                            "",
                        ),
    
                        followup_instructor=data.get(
                            "followup_instructor",
                            "",
                        ),
    
                        created_by=owner_id,
    
                    )
    
                    db.session.add(group_obj)
    
                    db.session.flush()
    
                    existing_groups[
                        group_number
                    ] = group_obj
    
                    result.created_groups += 1
    
                    result.group_count += 1
    
                # -------------------------------------------------------------
                # Construcción de datos del aprendiz
                # -------------------------------------------------------------
    
                # Solo actualizar campos realmente diligenciados.
                # Los campos ausentes en la plantilla se heredan del grupo o
                # conservan el valor existente del aprendiz.
                clean_data = {
                    key: value
                    for key, value in data.items()
                    if (
                        key in APPRENTICE_MODEL_FIELDS
                        and key != "group_id"
                        and value not in (None, "")
                    )
                }
    
                # -------------------------------------------------------------
                # Normalizar modalidad EP
                # -------------------------------------------------------------
    
                raw_ep = data.get("ep_modality")
    
                if raw_ep not in (None, ""):
                    canon_ep = canonical_ep_modality(raw_ep)
    
                    if canon_ep:
                        clean_data["ep_modality"] = canon_ep
                    else:
                        # EpModality es un catálogo estricto. Un valor no
                        # reconocido no debe llegar al modelo ni borrar un
                        # valor existente durante una actualización.
                        clean_data.pop("ep_modality", None)
                        current_app.logger.warning(
                            "Import: ep_modality no reconocido para aprendiz %s: %r. Se omitirá.",
                            data.get("document_number"),
                            raw_ep,
                        )
    
                clean_data["group_id"] = group_obj.id
    
                clean_data["group_number"] = (
                    group_obj.group_number
                )
    
                # -------------------------------------------------------------
                # Heredar información del grupo
                # -------------------------------------------------------------
    
                for inherited in (
    
                    "program_name",
    
                    "program_level",
    
                    "lead_instructor",
    
                    "followup_instructor",
    
                    "group_validity",
    
                ):
    
                    if not clean_data.get(inherited):
    
                        clean_data[inherited] = (
                            getattr(
                                group_obj,
                                inherited,
                                "",
                            )
                            or ""
                        )
    
                apprentice_obj = (
                    existing_apprentices.get(
                        data["document_number"]
                    )
                )
    
                if apprentice_obj is None:
    
                    apprentice_obj = Apprentice(
                        created_by=owner_id,
                        **clean_data,
                    )
    
                    db.session.add(
                        apprentice_obj
                    )
    
                    db.session.flush()
    
                    existing_apprentices[
                        data["document_number"]
                    ] = apprentice_obj
    
                    result.created_apprentices += 1
    
                else:
    
                    for key, value in clean_data.items():
    
                        if hasattr(
                            apprentice_obj,
                            key,
                        ):
                            setattr(
                                apprentice_obj,
                                key,
                                value,
                            )
    
                    apprentice_obj.created_by = (
                        owner_id
                    )
    
                    result.updated_apprentices += 1
    
                upsert_student_user(
                    apprentice_obj,
                    known_users=known_users,
                )
    
                ensure_submissions_for_apprentice(
                    apprentice_obj
                )
    
                result.apprentice_count += 1
    
        db.session.commit()
    
        return result
    except Exception:
        db.session.rollback()
        raise
